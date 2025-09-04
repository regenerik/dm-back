from flask import Blueprint, send_file, make_response, request, jsonify, render_template, current_app, Response # Blueprint para modularizar y relacionar con app
from flask_bcrypt import Bcrypt                                  # Bcrypt para encriptación
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity   # Jwt para tokens
from database import db                                          # importa la db desde database.py
from logging_config import logger
import os                                                        # Para datos .env
from dotenv import load_dotenv                                   # Para datos .env
load_dotenv()
from utils.data_mentor_utils import query_assistant_mentor
import urllib.request
import urllib.error
import json
import pandas as pd
from models import Usuarios_Por_Asignacion, Usuarios_Sin_ID, ValidaUsuarios,DetalleApies, AvanceCursada, DetallesDeCursos, CursadasAgrupadas,FormularioGestor,CuartoSurveySql, QuintoSurveySql, Comentarios2023, Comentarios2024, Comentarios2025, BaseLoopEstaciones, FichasGoogleCompetencia, FichasGoogle, SalesForce, ComentariosCompetencia, FileDailyID
import hashlib
from sqlalchemy.exc import SQLAlchemyError
import csv
import time
import tempfile
from openai import OpenAI
import httpx
from tempfile import NamedTemporaryFile
import openpyxl
from io import BytesIO
from datetime import datetime, timedelta 



OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")
if not OPENAI_API_KEY:
    raise ValueError("Debes definir la variable de entorno OPENAI_API_KEY con tu clave de API.")

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

HEADERS = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {OPENAI_API_KEY}",
    "OpenAI-Beta": "assistants=v2"
}


data_mentor_bp = Blueprint('data_mentor_bp', __name__)     # instanciar admin_bp desde clase Blueprint para crear las rutas.
bcrypt = Bcrypt()
jwt = JWTManager()

# Sistema de key base pre rutas ------------------------:

API_KEY = os.getenv('API_KEY')

def check_api_key(api_key):
    return api_key == API_KEY

@data_mentor_bp.before_request
def authorize():
    if request.method == 'OPTIONS':
        return
    if request.path in ['/horas-por-curso','/test_data_mentor_bp']:
        return
    api_key = request.headers.get('Authorization')
    if not api_key or not check_api_key(api_key):
        return jsonify({'message': 'Unauthorized'}), 401
    
# RUTA TEST:

@data_mentor_bp.route('/test_data_mentor_bp', methods=['GET'])
def test():
    logger.info("Chat data mentor bp rutas funcionando ok segun test.")
    return jsonify({'message': 'test bien sucedido','status':"Si lees esto, chat data mentor rutas funcionan bien..."}),200

@data_mentor_bp.route("/chat_mentor", methods=["POST"])
def chat_mentor():
    logger.info("1 - Entró en la ruta Chat_mentor.")
    """
    Recibe prompt y opcionalmente thread_id.
    """
    data = request.get_json()
    if not data or "prompt" not in data:
        return jsonify({"error": "Falta el prompt en el cuerpo de la solicitud"}), 400

    prompt = data["prompt"]
    thread_id = data.get("thread_id")  # puede ser None
    logger.info("2 - Encontró la data del prompt...")
    try:
        response_text, current_thread = query_assistant_mentor(prompt, thread_id)
        return jsonify({"response": response_text, "thread_id": current_thread}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    

@data_mentor_bp.route("/close_chat_mentor", methods=["POST"])
def close_chat():
    """
    Ruta para cerrar el thread del chat.
    Se espera recibir un JSON con la clave "thread_id".
    Llama al endpoint DELETE de la API para cerrar el hilo usando urllib.
    """
    data = request.get_json()
    if not data or "thread_id" not in data:
        return jsonify({"error": "Falta el thread_id en el cuerpo de la solicitud"}), 400

    thread_id = data["thread_id"]
    delete_url = f"https://api.openai.com/v1/threads/{thread_id}"

    try:
        req = urllib.request.Request(delete_url, headers=HEADERS, method="DELETE")
        with urllib.request.urlopen(req) as response:
            result_data = response.read().decode("utf-8")
            result = json.loads(result_data)
        return jsonify(result), 200
    except urllib.error.HTTPError as e:
        error_message = e.read().decode("utf-8")
        return jsonify({"error": f"HTTPError {e.code}: {error_message}"}), e.code
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@data_mentor_bp.route('/horas-por-curso', methods=['GET'])
def horas_por_curso():
    data = [
        {"curso": "Node.js Básico", "horas": 5},
        {"curso": "React Intermedio", "horas": 7},
        {"curso": "Flask Fullstack", "horas": 9}
    ]
    return jsonify(data)

# -------------------------- MODELOS QUE SE TIENEN EN CUENTA PARA CONTABILIZAR SUS REGISTROS ----------------------------------

# Diccionario para mapear nombres de string a clases reales
MODELS = {
    'Usuarios_Por_Asignacion': Usuarios_Por_Asignacion,
    'Usuarios_Sin_ID': Usuarios_Sin_ID,
    'ValidaUsuarios': ValidaUsuarios,
    'DetalleApies' : DetalleApies,
    'AvanceCursada': AvanceCursada,
    'DetallesDeCursos' : DetallesDeCursos,
    'CursadasAgrupadas' : CursadasAgrupadas,
    'FormularioGestor' :FormularioGestor,
    'CuartoSurveySql': CuartoSurveySql,
    'QuintoSurveySql' : QuintoSurveySql,
    'Comentarios2023': Comentarios2023,
    'Comentarios2024': Comentarios2024,
    'Comentarios2025': Comentarios2025,
    'BaseLoopEstaciones' : BaseLoopEstaciones,
    'FichasGoogleCompetencia' : FichasGoogleCompetencia,
    'FichasGoogle' : FichasGoogle,
    'SalesForce' : SalesForce,
    'ComentariosCompetencia' : ComentariosCompetencia
    # Agregá los modelos que quieras habilitar acá
}

# -------------------------- Contabilizar longitud de cualquier tabla ----------------------------------

@data_mentor_bp.route('/contar-registros', methods=['POST'])
def contar_registros():
    data = request.get_json()
    nombre_tabla = data.get('tabla')

    if not nombre_tabla:
        return jsonify({"error": "Falta el nombre de la tabla"}), 400

    modelo = MODELS.get(nombre_tabla)
    if not modelo:
        return jsonify({"error": f"La tabla '{nombre_tabla}' no está habilitada o no existe"}), 404

    try:
        cantidad = db.session.query(modelo).count()
        return jsonify({"tabla": nombre_tabla, "cantidad_registros": cantidad}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# -------------------------- ACA VIENEN LAS RUTAS DE LAS TABLAS DE REPORTES ----------------------------


@data_mentor_bp.route('/usuarios_por_asignacion/<int:registro_id>', methods=['GET'])
def get_usuario_por_asignacion(registro_id):
    """
    Devuelve el registro de Usuarios_Por_Asignacion con el id dado,
    usando el método serialize() del modelo.
    """
    logger.info(f"Buscando Usuarios_Por_Asignacion id={registro_id}")
    registro = Usuarios_Por_Asignacion.query.get(registro_id)

    if not registro:
        logger.warning(f"Usuarios_Por_Asignacion id={registro_id} no encontrado")
        return jsonify({
            'error': 'Registro no encontrado',
            'status': 404
        }), 404

    logger.info(f"Registro encontrado: {registro}")
    return jsonify(registro.serialize()), 200

@data_mentor_bp.route('/usuarios_sin_id/<int:registro_id>', methods=['GET'])
def get_usuario_sin_id(registro_id):
    """
    Devuelve el registro de Usuarios_Sin_ID con el id dado,
    usando el método serialize() del modelo.
    """
    logger.info(f"Buscando Usuarios_Sin_ID id={registro_id}")
    registro = Usuarios_Sin_ID.query.get(registro_id)

    if not registro:
        logger.warning(f"Usuarios_Sin_ID id={registro_id} no encontrado")
        return jsonify({
            'error': 'Registro no encontrado',
            'status': 404
        }), 404

    logger.info(f"Registro encontrado: {registro}")
    return jsonify(registro.serialize()), 200




# RUTAS PARA CARGAR TABLAS DE EXPERIENCIA 2023 24 y 25

@data_mentor_bp.route('/cargar_comentarios_2023', methods=['POST'])
def cargar_comentarios_encuesta_2023():
    """
    Recibe un archivo .xlsx vía form-data (campo: 'file') y guarda sus registros en la DB
    """
    archivo = request.files.get('file')
    if not archivo:
        return jsonify({'error': 'No se envió ningún archivo', 'status': 400}), 400

    try:
        df = pd.read_excel(archivo)

        registros = []
        for _, fila in df.iterrows():
            fecha_raw = fila.get('FECHA')
            try:
                fecha = pd.to_datetime(fecha_raw) if pd.notnull(fecha_raw) else None
            except:
                fecha = None

            nuevo = Comentarios2023(
                fecha=fecha,
                apies=str(fila.get('APIES', '')).strip(),
                comentario=str(fila.get('COMENTARIO', '')).strip(),
                canal=str(fila.get('CANAL', '')).strip(),
                topico=str(fila.get('TÓPICO', '')).strip(),
                sentiment=str(fila.get('SENTIMENT', '')).strip()
            )
            registros.append(nuevo)

        db.session.add_all(registros)
        db.session.commit()

        return jsonify({'mensaje': f'Se guardaron {len(registros)} comentarios', 'status': 200}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}', 'status': 500}), 500
    
@data_mentor_bp.route('/cargar_comentarios_2024', methods=['POST'])
def cargar_comentarios_encuesta_2024():
    """
    Recibe un archivo .xlsx vía form-data (campo: 'file') y guarda sus registros en la DB.
    Optimizado para bajo uso de memoria mediante procesamiento fila a fila con openpyxl.
    Incluye logs de progreso consolidados y tiempo de ejecución en la respuesta.
    """
    start_time = datetime.now() # Iniciar el temporizador

    archivo = request.files.get('file')
    if not archivo:
        logger.error("No se envió ningún archivo en la solicitud.")
        return jsonify({'error': 'No se envió ningún archivo', 'status': 400}), 400

    total_registros_guardados = 0
    BATCH_SIZE = 5000 
    
    # Contadores y muestras para errores de fecha
    date_errors_count = 0
    sample_date_errors = []
    MAX_SAMPLE_DATE_ERRORS = 5 # Cuántos ejemplos de errores de fecha guardar

    logger.info("======================================================")
    logger.info("============= INICIANDO PROCESO DE CARGA ==============")
    logger.info("======================================================")
    logger.info(f"Hora de inicio: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"Archivo recibido: '{archivo.filename}' ({archivo.content_length / (1024*1024):.2f} MB)")
    logger.info(f"Tamaño de lote (BATCH_SIZE) para DB: {BATCH_SIZE} filas por commit.")

    try:
        excel_data = BytesIO(archivo.read())
        workbook = openpyxl.load_workbook(excel_data, read_only=True, data_only=True)
        
        if not workbook.sheetnames:
            logger.error("El archivo Excel recibido no contiene hojas.")
            return jsonify({'error': 'El archivo Excel no contiene hojas.', 'status': 400}), 400
        
        sheet = workbook[workbook.sheetnames[0]] 
        logger.info(f"Se procesará la hoja: '{sheet.title}'")

        # Asumimos que la primera fila son los encabezados
        headers = [cell.value for cell in sheet[1]] 
        header_map = {str(header).strip().upper(): i for i, header in enumerate(headers) if header}
        
        col_fecha = header_map.get('FECHA')
        col_apies = header_map.get('APIES')
        col_comentario = header_map.get('COMENTARIO')
        col_canal = header_map.get('CANAL')
        col_topico = header_map.get('TÓPICO')
        col_sentiment = header_map.get('SENTIMENT')

        if col_fecha is None or col_apies is None or col_comentario is None or \
           col_canal is None or col_topico is None or col_sentiment is None:
            missing_cols = []
            if col_fecha is None: missing_cols.append('FECHA')
            if col_apies is None: missing_cols.append('APIES')
            if col_comentario is None: missing_cols.append('COMENTARIO')
            if col_canal is None: missing_cols.append('CANAL')
            if col_topico is None: missing_cols.append('TÓPICO')
            if col_sentiment is None: missing_cols.append('SENTIMENT')
            
            error_msg = f"Faltan columnas requeridas en el archivo Excel: {', '.join(missing_cols)}. Asegúrate de que los encabezados coincidan (MAYÚSCULAS)."
            logger.error(error_msg)
            return jsonify({'error': error_msg, 'status': 400}), 400

        registros_chunk = []
        row_counter_total = 0 
        chunk_counter = 0

        # Iterar sobre las filas de la hoja, empezando desde la segunda fila
        # enumerate(..., start=2) para obtener el número de fila real del Excel (incluyendo encabezado)
        for row_index_excel, row in enumerate(sheet.iter_rows(min_row=2), start=2): 
            row_counter_total += 1
            
            fecha_raw = row[col_fecha].value if col_fecha is not None else None
            apies = row[col_apies].value if col_apies is not None else ''
            comentario = row[col_comentario].value if col_comentario is not None else ''
            canal = row[col_canal].value if col_canal is not None else ''
            topico = row[col_topico].value if col_topico is not None else ''
            sentiment = row[col_sentiment].value if col_sentiment is not None else ''
            
            fecha = None # Inicializar fecha como None
            try:
                if isinstance(fecha_raw, datetime):
                    fecha = fecha_raw # openpyxl ya lo parseó a datetime
                elif pd.notnull(fecha_raw) and str(fecha_raw).strip(): # Si tiene valor no nulo y no es string vacío
                    # Intentar parsear si no es datetime
                    fecha = pd.to_datetime(fecha_raw, errors='coerce') 
                # Si pd.to_datetime con errors='coerce' falla, fecha será NaT, que se convierte en None al asignar
            except Exception as date_e:
                # Solo logear si es un error que no sea solo 'coerce' a NaT, o si queremos más detalle
                pass # No loguear cada error de fecha individualmente
            
            # Contar errores de fecha y guardar una muestra
            if fecha is pd.NaT: # pd.NaT indica que pd.to_datetime falló y coecionó
                date_errors_count += 1
                if len(sample_date_errors) < MAX_SAMPLE_DATE_ERRORS:
                    sample_date_errors.append(f"Fila {row_index_excel}: '{fecha_raw}'")
                fecha = None # Asegurarse de que sea None para la DB si falló el parseo

            nuevo = Comentarios2024(
                fecha=fecha,
                apies=str(apies).strip(),
                comentario=str(comentario).strip(),
                canal=str(canal).strip(),
                topico=str(topico).strip(),
                sentiment=str(sentiment).strip()
            )
            registros_chunk.append(nuevo)

            # Si el chunk alcanza el tamaño definido, lo guardamos en la DB
            if len(registros_chunk) >= BATCH_SIZE:
                chunk_counter += 1
                logger.info(f"--- Procesando CHUNK {chunk_counter} (Filas Excel {row_counter_total - len(registros_chunk) + 1} - {row_counter_total}) ---")
                
                db.session.add_all(registros_chunk)
                db.session.commit()
                total_registros_guardados += len(registros_chunk)
                logger.info(f"CHUNK {chunk_counter} COMPLETO. Guardados {len(registros_chunk)} registros en DB. Total acumulado: {total_registros_guardados}")
                
                db.session.expunge_all() 
                del registros_chunk 
                registros_chunk = [] # Reiniciar la lista para el siguiente lote
                
                logger.info(f"--- Memoria del CHUNK {chunk_counter} liberada. ---")

        # Guardar los registros restantes (el último lote)
        if registros_chunk:
            chunk_counter += 1
            logger.info(f"--- Procesando CHUNK FINAL {chunk_counter} (Filas Excel {row_counter_total - len(registros_chunk) + 1} - {row_counter_total}) ---")
            db.session.add_all(registros_chunk)
            db.session.commit()
            total_registros_guardados += len(registros_chunk)
            logger.info(f"CHUNK FINAL {chunk_counter} COMPLETO. Guardados {len(registros_chunk)} registros en DB.")
            logger.info(f"TOTAL DE REGISTROS GUARDADOS EN LA DB: {total_registros_guardados}")
            db.session.expunge_all()
            del registros_chunk

        end_time = datetime.now() # Finalizar el temporizador
        time_elapsed = end_time - start_time
        # Formatear el tiempo de ejecución a HH:MM:SS
        seconds = int(time_elapsed.total_seconds())
        time_format = str(timedelta(seconds=seconds))

        logger.info("======================================================")
        logger.info("============ PROCESO DE CARGA FINALIZADO ============")
        logger.info(f"Hora de finalización: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"Tiempo total de ejecución: {time_format}")
        logger.info(f"TOTAL DE FILAS PROCESADAS EN EXCEL: {row_counter_total}")
        logger.info(f"TOTAL DE REGISTROS GUARDADOS EN LA DB: {total_registros_guardados}")
        if date_errors_count > 0:
            logger.warning(f"ADVERTENCIA: Se encontraron {date_errors_count} errores al parsear fechas. Ejemplos: {'; '.join(sample_date_errors)}")
        logger.info("======================================================")
        
        return jsonify({
            'mensaje': f'Se guardaron {total_registros_guardados} comentarios en total.',
            'status': 200,
            'tiempo_de_guardado': time_format,
            'detalles_procesamiento': {
                'total_filas_excel_leidas': row_counter_total,
                'total_registros_db_guardados': total_registros_guardados,
                'errores_fecha_contados': date_errors_count,
                'ejemplos_errores_fecha': sample_date_errors if date_errors_count > 0 else "Ninguno"
            }
        }), 200

    except Exception as e:
        db.session.rollback() 
        end_time = datetime.now() # Capturar tiempo de finalización incluso en error
        time_elapsed_on_error = end_time - start_time
        seconds_on_error = int(time_elapsed_on_error.total_seconds())
        time_format_on_error = str(timedelta(seconds=seconds_on_error))

        logger.error("======================================================")
        logger.error("============ ERROR FATAL DURANTE LA CARGA ============")
        logger.error(f"Hora del error: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        logger.error(f"Tiempo transcurrido hasta el error: {time_format_on_error}")
        logger.error(f"Error: {str(e)}", exc_info=True)
        logger.error("======================================================")
        return jsonify({
            'error': f'Error al procesar el archivo: {str(e)}', 
            'status': 500,
            'tiempo_transcurrido_hasta_error': time_format_on_error
        }), 500
    
@data_mentor_bp.route('/cargar_comentarios_2025', methods=['POST'])
def cargar_comentarios_encuesta_2025():
    """
    Recibe un archivo .xlsx vía form-data (campo: 'file') y guarda sus registros en la DB.
    Implementa lógica incremental (deduplicación por hash_id) y optimización de memoria
    mediante openpyxl y procesamiento por lotes.
    Incluye logs detallados de progreso y tiempo de ejecución.
    """
    start_time = datetime.now() # Iniciar el temporizador

    archivo = request.files.get('file')
    if not archivo:
        logger.error("No se envió ningún archivo en la solicitud.")
        return jsonify({'error': 'No se envió ningún archivo', 'status': 400}), 400

    # Contadores para el resumen final
    total_registros_guardados = 0
    total_filas_excel_leidas = 0
    date_errors_count = 0
    sample_date_errors = []
    MAX_SAMPLE_DATE_ERRORS = 5 
    duplicados_en_archivo_contados = 0
    duplicados_en_db_contados = 0

    # Este set rastreará todos los hashes únicos encontrados en el ARCHIVO ACTUAL
    # para evitar duplicados que aparezcan en diferentes lotes del mismo archivo.
    all_hashes_seen_in_current_file_session = set() 

    BATCH_SIZE = 5000 

    logger.info("======================================================")
    logger.info("====== INICIANDO PROCESO DE CARGA COMENTARIOS 2025 ======")
    logger.info("======================================================")
    logger.info(f"Hora de inicio: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"Archivo recibido: '{archivo.filename}' ({archivo.content_length / (1024*1024):.2f} MB)")
    logger.info(f"Tamaño de lote (BATCH_SIZE) para DB: {BATCH_SIZE} filas por commit.")

    try:
        excel_data = BytesIO(archivo.read())
        workbook = openpyxl.load_workbook(excel_data, read_only=True, data_only=True)
        
        if not workbook.sheetnames:
            logger.error("El archivo Excel recibido no contiene hojas.")
            return jsonify({'error': 'El archivo Excel no contiene hojas.', 'status': 400}), 400
        
        sheet = workbook[workbook.sheetnames[0]] 
        logger.info(f"Se procesará la hoja: '{sheet.title}'")

        # Asumimos que la primera fila son los encabezados
        headers = [cell.value for cell in sheet[1]] 
        header_map = {str(header).strip().upper(): i for i, header in enumerate(headers) if header}
        
        # Mapeo de columnas, asegurando que existan
        col_fecha = header_map.get('FECHA')
        col_apies = header_map.get('APIES')
        col_comentario = header_map.get('COMENTARIO')
        col_canal = header_map.get('CANAL')
        col_topico = header_map.get('TÓPICO')
        col_sentiment = header_map.get('SENTIMENT')

        # Verificar que las columnas obligatorias existan
        required_cols = {'FECHA', 'APIES', 'COMENTARIO', 'CANAL', 'TÓPICO', 'SENTIMENT'}
        missing_cols = [col for col in required_cols if header_map.get(col) is None]
        if missing_cols:
            error_msg = f"Faltan columnas requeridas en el archivo Excel: {', '.join(missing_cols)}. Asegúrate de que los encabezados coincidan (MAYÚSCULAS)."
            logger.error(error_msg)
            return jsonify({'error': error_msg, 'status': 400}), 400

        registros_para_lote_db = [] # Esta lista contendrá los objetos ORM para el bulk_save_objects
        chunk_counter = 0

        # Iterar sobre las filas de la hoja, empezando desde la segunda fila (después del encabezado)
        for row_index_excel, row in enumerate(sheet.iter_rows(min_row=2), start=2): 
            total_filas_excel_leidas += 1
            
            # Extraer valores de celda de forma segura
            fecha_raw = row[col_fecha].value
            apies = row[col_apies].value
            comentario = row[col_comentario].value
            canal = row[col_canal].value
            topico = row[col_topico].value
            sentiment = row[col_sentiment].value
            
            # --- Manejo de la fecha y generación del hash ---
            fecha = None
            try:
                if isinstance(fecha_raw, datetime):
                    fecha = fecha_raw 
                elif pd.notnull(fecha_raw) and str(fecha_raw).strip():
                    fecha = pd.to_datetime(fecha_raw, errors='coerce') 
                
                if fecha is pd.NaT: # Si el parseo de Pandas falló y coecionó
                    date_errors_count += 1
                    if len(sample_date_errors) < MAX_SAMPLE_DATE_ERRORS:
                        sample_date_errors.append(f"Fila {row_index_excel}: '{fecha_raw}'")
                    fecha = None # Asegurarse de que sea None para la DB si falló el parseo
            except Exception as date_e:
                date_errors_count += 1
                if len(sample_date_errors) < MAX_SAMPLE_DATE_ERRORS:
                    sample_date_errors.append(f"Fila {row_index_excel}: '{fecha_raw}' - {date_e}")
                fecha = None 
            
            # Stripear strings antes de generar hash y crear objeto
            apies = str(apies if apies is not None else '').strip()
            comentario = str(comentario if comentario is not None else '').strip()
            canal = str(canal if canal is not None else '').strip()
            topico = str(topico if topico is not None else '').strip()
            sentiment = str(sentiment if sentiment is not None else '').strip()

            # Generar hash único para este registro
            hash_id = Comentarios2025.generar_hash(fecha, apies, comentario, canal, topico, sentiment)

            # Verificar si este hash_id ya fue visto en el archivo ACTUAL (en cualquier lote procesado)
            if hash_id in all_hashes_seen_in_current_file_session:
                duplicados_en_archivo_contados += 1
                # logger.debug(f"Saltando duplicado DENTRO DEL ARCHIVO (hash: {hash_id}) en fila {row_index_excel}.")
                continue # Saltar esta fila, ya fue procesada o es un duplicado interno

            all_hashes_seen_in_current_file_session.add(hash_id) # Registrar este hash_id como visto

            comentario_obj = Comentarios2025(
                fecha=fecha,
                apies=apies,
                comentario=comentario,
                canal=canal,
                topico=topico,
                sentiment=sentiment,
                hash_id=hash_id # Asignar el hash_id al objeto
            )
            registros_para_lote_db.append(comentario_obj)

            # Si el lote para DB alcanza el tamaño definido, procesamos y guardamos
            if len(registros_para_lote_db) >= BATCH_SIZE:
                chunk_counter += 1
                logger.info(f"--- Procesando LOTE {chunk_counter} ({len(registros_para_lote_db)} candidatos para DB) ---")
                
                # Obtener los hashes de este lote para consultar la DB
                hashes_en_lote = [obj.hash_id for obj in registros_para_lote_db]
                
                # Consultar la DB para ver cuáles de estos hashes ya existen
                existentes_en_db_lote = set(
                    r[0] for r in db.session.query(Comentarios2025.hash_id)
                    .filter(Comentarios2025.hash_id.in_(hashes_en_lote))
                    .all()
                )
                
                # Filtrar los que realmente son nuevos para insertar
                nuevos_para_insertar = [
                    obj for obj in registros_para_lote_db 
                    if obj.hash_id not in existentes_en_db_lote
                ]
                
                duplicados_en_db_contados += (len(registros_para_lote_db) - len(nuevos_para_insertar))

                if nuevos_para_insertar:
                    # Usamos bulk_save_objects para una inserción más eficiente
                    db.session.bulk_save_objects(nuevos_para_insertar)
                    db.session.commit()
                    total_registros_guardados += len(nuevos_para_insertar)
                    logger.info(f"LOTE {chunk_counter} COMPLETO. Insertados {len(nuevos_para_insertar)} nuevos registros en DB.")
                    logger.info(f"TOTAL ACUMULADO GUARDADOS: {total_registros_guardados}")
                    logger.info(f"TOTAL ACUMULADO DUPLICADOS EN DB IGNORADOS: {duplicados_en_db_contados}")
                else:
                    logger.info(f"LOTE {chunk_counter} COMPLETO. No se encontraron registros nuevos para insertar en DB.")
                
                db.session.expunge_all() 
                del registros_para_lote_db 
                registros_para_lote_db = [] # Reiniciar la lista para el siguiente lote
                
                logger.info(f"--- Memoria del LOTE {chunk_counter} liberada. ---")

        # --- Fin del bucle de filas ---

        # Procesar y guardar los registros restantes (el último lote, si lo hay)
        if registros_para_lote_db:
            chunk_counter += 1
            logger.info(f"--- Procesando LOTE FINAL {chunk_counter} ({len(registros_para_lote_db)} candidatos restantes) ---")
            
            hashes_en_lote = [obj.hash_id for obj in registros_para_lote_db]
            existentes_en_db_lote = set(
                r[0] for r in db.session.query(Comentarios2025.hash_id)
                .filter(Comentarios2025.hash_id.in_(hashes_en_lote))
                .all()
            )
            nuevos_para_insertar = [
                obj for obj in registros_para_lote_db 
                if obj.hash_id not in existentes_en_db_lote
            ]
            duplicados_en_db_contados += (len(registros_para_lote_db) - len(nuevos_para_insertar))

            if nuevos_para_insertar:
                db.session.bulk_save_objects(nuevos_para_insertar)
                db.session.commit()
                total_registros_guardados += len(nuevos_para_insertar)
                logger.info(f"LOTE FINAL {chunk_counter} COMPLETO. Insertados {len(nuevos_para_insertar)} nuevos registros en DB.")
            else:
                logger.info(f"LOTE FINAL {chunk_counter} COMPLETO. No se encontraron registros nuevos para insertar en DB.")
            
            db.session.expunge_all()
            del registros_para_lote_db

        end_time = datetime.now() 
        time_elapsed = end_time - start_time
        seconds = int(time_elapsed.total_seconds())
        time_format = str(timedelta(seconds=seconds))

        logger.info("======================================================")
        logger.info("========== PROCESO DE CARGA COMENTARIOS 2025 FINALIZADO ==========")
        logger.info(f"Hora de finalización: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"Tiempo total de ejecución: {time_format}")
        logger.info(f"TOTAL DE FILAS LEÍDAS DEL EXCEL: {total_filas_excel_leidas}")
        logger.info(f"TOTAL DE REGISTROS NUEVOS GUARDADOS EN LA DB: {total_registros_guardados}")
        logger.info(f"TOTAL DE DUPLICADOS EN EL ARCHIVO IGNORADOS: {duplicados_en_archivo_contados}")
        logger.info(f"TOTAL DE DUPLICADOS EN LA DB IGNORADOS: {duplicados_en_db_contados}")
        if date_errors_count > 0:
            logger.warning(f"ADVERTENCIA: Se encontraron {date_errors_count} errores al parsear fechas. Ejemplos: {'; '.join(sample_date_errors)}")
        logger.info("======================================================")
        
        return jsonify({
            'mensaje': f'Se guardaron {total_registros_guardados} comentarios nuevos en total.',
            'status': 200,
            'tiempo_de_guardado': time_format,
            'detalles_procesamiento': {
                'total_filas_excel_leidas': total_filas_excel_leidas,
                'total_registros_db_guardados': total_registros_guardados,
                'duplicados_ignorados_en_archivo': duplicados_en_archivo_contados,
                'duplicados_ignorados_en_db': duplicados_en_db_contados,
                'errores_fecha_contados': date_errors_count,
                'ejemplos_errores_fecha': sample_date_errors if date_errors_count > 0 else "Ninguno"
            }
        }), 200

    except Exception as e:
        db.session.rollback() 
        end_time = datetime.now() 
        time_elapsed_on_error = end_time - start_time
        seconds_on_error = int(time_elapsed_on_error.total_seconds())
        time_format_on_error = str(timedelta(seconds=seconds_on_error))

        logger.error("======================================================")
        logger.error("============ ERROR FATAL DURANTE LA CARGA ============")
        logger.error(f"Hora del error: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        logger.error(f"Tiempo transcurrido hasta el error: {time_format_on_error}")
        logger.error(f"Error: {str(e)}", exc_info=True)
        logger.error("======================================================")
        return jsonify({
            'error': f'Error al procesar el archivo: {str(e)}', 
            'status': 500,
            'tiempo_transcurrido_hasta_error': time_format_on_error,
            'detalles_procesamiento': {
                'total_filas_excel_leidas': total_filas_excel_leidas,
                'total_registros_db_guardados_hasta_error': total_registros_guardados # Registros guardados antes del error
            }
        }), 500
    
@data_mentor_bp.route('/cargar_base_loop', methods=['POST'])
def cargar_base_loop():
    archivo = request.files.get('file')
    if not archivo:
        return jsonify({"error": "No se envió ningún archivo"}), 400

    try:
        # Detectar delimitador (coma o punto y coma)
        sample = archivo.read(2048).decode('utf-8')
        archivo.seek(0)
        delimiter = csv.Sniffer().sniff(sample).delimiter

        df = pd.read_csv(archivo, sep=delimiter, encoding='utf-8', on_bad_lines='skip')
    except Exception as e:
        return jsonify({"error": f"Error al leer el archivo CSV: {e}"}), 400

    try:
        columnas_db = {col.name: col.key for col in BaseLoopEstaciones.__table__.columns}

        registros = []
        for _, fila in df.iterrows():
            datos_instancia = {}
            for nombre_col_csv, valor in fila.items():
                if nombre_col_csv in columnas_db:
                    campo_para_modelo = columnas_db[nombre_col_csv]
                    datos_instancia[campo_para_modelo] = valor

            # Instanciamos así para evitar el error de keyword inválido
            instancia = BaseLoopEstaciones()
            for key, val in datos_instancia.items():
                setattr(instancia, key, val)
            registros.append(instancia)

        db.session.bulk_save_objects(registros)
        db.session.commit()

        return jsonify({"mensaje": f"{len(registros)} registros insertados correctamente"}), 200

    except SQLAlchemyError as e:
        db.session.rollback()
        return jsonify({"error": f"Error al insertar en la base de datos: {e}"}), 500
    except Exception as e:
        return jsonify({"error": f"Error al procesar el archivo: {e}"}), 500
    
@data_mentor_bp.route('/cargar_fichas_google_competencia', methods=['POST'])
def cargar_fichas_google_competencia():
    archivo = request.files.get('file')
    if not archivo:
        return jsonify({'error': 'No se envió ningún archivo', 'status': 400}), 400

    try:
        df = pd.read_excel(archivo)

        candidatos = []
        hash_ids = []
        hash_set_memoria = set()

        for _, fila in df.iterrows():
            id_loop = str(fila.get('idLoop', '')).strip()
            total_review_count = str(fila.get('totalReviewCount', '')).strip()
            average_rating = str(fila.get('averageRating', '')).strip()

            # Saltear filas con campos vacíos importantes
            if not id_loop or not total_review_count or not average_rating:
                continue

            hash_id = FichasGoogleCompetencia.generar_hash(id_loop, total_review_count, average_rating)

            # Saltear si ya está en memoria (repetido en el mismo archivo)
            if hash_id in hash_set_memoria:
                continue
            hash_set_memoria.add(hash_id)

            ficha_obj = FichasGoogleCompetencia(
                id_loop=id_loop,
                total_review_count=total_review_count,
                average_rating=average_rating,
                hash_id=hash_id
            )

            candidatos.append(ficha_obj)
            hash_ids.append(hash_id)

        # Buscar duplicados ya existentes en la DB
        existentes = set(
            r[0] for r in db.session.query(FichasGoogleCompetencia.hash_id)
            .filter(FichasGoogleCompetencia.hash_id.in_(hash_ids))
            .all()
        )

        nuevos = [f for f in candidatos if f.hash_id not in existentes]

        if nuevos:
            db.session.bulk_save_objects(nuevos)
            db.session.commit()

        return jsonify({
            'mensaje': f'Se guardaron {len(nuevos)} fichas nuevas',
            'preexistentes_ignorados': len(candidatos) - len(nuevos),
            'status': 200
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}', 'status': 500}), 500

@data_mentor_bp.route('/cargar_fichas_google', methods=['POST'])
def cargar_fichas_google():
    archivo = request.files.get('file')
    if not archivo:
        return jsonify({'error': 'No se envió ningún archivo', 'status': 400}), 400

    try:
        df = pd.read_excel(archivo)

        candidatos = []
        hash_ids = []
        hash_set_memoria = set()

        for _, fila in df.iterrows():
            store_code = str(fila.get('Store Code', '')).strip()
            cantidad_de_calificaciones = str(fila.get('Cantidad de calificaciones', '')).strip()
            start_rating = str(fila.get('Star Rating', '')).strip()

            # Saltear filas con campos vacíos importantes
            if not store_code or not cantidad_de_calificaciones or not start_rating:
                continue

            hash_id = FichasGoogle.generar_hash(store_code, cantidad_de_calificaciones, start_rating)

            # Saltear si ya está en memoria (repetido en el mismo archivo)
            if hash_id in hash_set_memoria:
                continue
            hash_set_memoria.add(hash_id)

            ficha_obj = FichasGoogle(
                store_code=store_code,
                cantidad_de_calificaciones=cantidad_de_calificaciones,
                start_rating=start_rating,
                hash_id=hash_id
            )

            candidatos.append(ficha_obj)
            hash_ids.append(hash_id)

        # Buscar duplicados ya existentes en la DB
        existentes = set(
            r[0] for r in db.session.query(FichasGoogle.hash_id)
            .filter(FichasGoogle.hash_id.in_(hash_ids))
            .all()
        )

        nuevos = [f for f in candidatos if f.hash_id not in existentes]

        if nuevos:
            db.session.bulk_save_objects(nuevos)
            db.session.commit()

        return jsonify({
            'mensaje': f'Se guardaron {len(nuevos)} fichas nuevas',
            'preexistentes_ignorados': len(candidatos) - len(nuevos),
            'status': 200
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}', 'status': 500}), 500
    
@data_mentor_bp.route('/cargar_salesforce', methods=['POST'])
def cargar_salesforce():
    archivo = request.files.get('file')
    if not archivo:
        return jsonify({'error': 'No se envió ningún archivo', 'status': 400}), 400

    try:
        df = pd.read_excel(archivo)

        candidatos = []
        hash_ids = []
        hash_set_memoria = set()

        for _, fila in df.iterrows():
            valores = [
                fila.get('Estacion de Servicio: Zona', ''),
                fila.get('Número del caso', ''),
                fila.get('Estado', ''),
                fila.get('Tipificación Caso', ''),
                fila.get('Asunto', ''),
                fila.get('Fecha/Hora de apertura', ''),
                fila.get('Cantidad de Reclamos', ''),
                fila.get('Defensa al Consumidor', ''),
                fila.get('GGRR/COLA Asignado', ''),
                fila.get('Propietario del caso: Nombre completo', ''),
                fila.get('Descripción', ''),
                fila.get('Nombre del contacto: Nombre completo', ''),
                fila.get('Comentarios', ''),
                fila.get('Estacion de Servicio: Razón Social', ''),
                fila.get('Estacion de Servicio: Red', ''),
                fila.get('Estacion de Servicio: Regional', ''),
            ]

            hash_id = SalesForce.generar_hash(*valores)

            if hash_id in hash_set_memoria:
                continue
            hash_set_memoria.add(hash_id)

            registro = SalesForce(
                estacion_servicio_zona=valores[0],
                numero_de_caso=valores[1],
                estado=valores[2],
                tipificacion_caso=valores[3],
                asunto=valores[4],
                fecha_apertura=valores[5],
                cantidad_reclamos=valores[6],
                defensa_consumidor=valores[7],
                ggrr_cola_asignado=valores[8],
                propietario_nombre=valores[9],
                descripcion=valores[10],
                contacto_nombre=valores[11],
                comentarios=valores[12],
                razon_social=valores[13],
                red=valores[14],
                regional=valores[15],
                hash_id=hash_id
            )

            candidatos.append(registro)
            hash_ids.append(hash_id)

        existentes = set(
            r[0] for r in db.session.query(SalesForce.hash_id)
            .filter(SalesForce.hash_id.in_(hash_ids))
            .all()
        )

        nuevos = [r for r in candidatos if r.hash_id not in existentes]

        if nuevos:
            db.session.bulk_save_objects(nuevos)
            db.session.commit()

        return jsonify({
            'mensaje': f'Se guardaron {len(nuevos)} casos nuevos',
            'preexistentes_ignorados': len(candidatos) - len(nuevos),
            'status': 200
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}', 'status': 500}), 500
    
@data_mentor_bp.route('/cargar_comentarios_competencia', methods=['POST'])
def cargar_comentarios_competencia():
    archivo = request.files.get('file')
    if not archivo:
        return jsonify({'error': 'No se recibió ningún archivo'}), 400

    try:
        df = pd.read_excel(archivo)

        # Normalizar la fecha
        if 'FECHA' in df.columns:
            df['FECHA'] = df['FECHA'].astype(str)


        # Renombrar columna ID si existe
        if 'ID' in df.columns:
            df.rename(columns={'ID': 'ID_ORIGINAL'}, inplace=True)

        # Reemplazamos espacios y ponemos mayúsculas para asegurar
        df.columns = [col.upper().replace(" ", "_") for col in df.columns]

        nuevos = []
        for _, fila in df.iterrows():
            hash_id = ComentariosCompetencia.generar_hash(
                fila.get('ID_ORIGINAL', ''),
                fila.get('FECHA', ''),
                fila.get('IDLOOP', ''),
                fila.get('COMENTARIO', ''),
                fila.get('RATING', ''),
                fila.get('SENTIMIENTO', ''),
                fila.get('TÓPICO', '')
            )

            # Chequeamos si ya existe
            if not ComentariosCompetencia.query.filter_by(hash_id=hash_id).first():
                nuevo = ComentariosCompetencia(
                    id_original=fila.get('ID_ORIGINAL'),
                    fecha=fila.get('FECHA'),
                    id_loop=fila.get('IDLOOP'),
                    comentario=fila.get('COMENTARIO'),
                    rating=fila.get('RATING'),
                    sentimiento=fila.get('SENTIMIENTO'),
                    topico=fila.get('TÓPICO'),
                    hash_id=hash_id
                )
                nuevos.append(nuevo)

        db.session.bulk_save_objects(nuevos)
        db.session.commit()

        return jsonify({'guardados': len(nuevos)})

    except Exception as e:
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}'}), 500
    


# LA COMPILACION Y SUBIDA DEL JSON DE TODA LA DATA>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
# --- Funciones auxiliares para el proceso asíncrono (mantienen la lógica previa) ---
from concurrent.futures import ThreadPoolExecutor
executor = ThreadPoolExecutor(max_workers=5)
def _wait_for_file_processing(file_id: str):
    start_time = time.time()
    max_wait_time = 300 
    while True:
        try:
            file_obj = client.files.retrieve(file_id=file_id)
            if file_obj.status == "processed":
                logger.info(f"Archivo {file_id} procesado por OpenAI en {time.time() - start_time:.2f} segundos.")
                return file_obj
            elif file_obj.status == "failed":
                error_message = f"El procesamiento del archivo {file_id} falló en OpenAI. Detalles: {file_obj.last_error.message if file_obj.last_error else 'Desconocido'}. Por favor, revise el archivo y el log de OpenAI."
                logger.error(error_message)
                raise Exception(error_message) 
            else:
                current_wait_time = time.time() - start_time
                if current_wait_time > max_wait_time:
                    error_message = f"Tiempo de espera excedido ({max_wait_time}s) para el procesamiento del archivo {file_id}. Estado actual: {file_obj.status}"
                    logger.error(error_message)
                    raise TimeoutError(error_message) 
                logger.info(f"Archivo {file_id} en estado '{file_obj.status}', esperando... ({current_wait_time:.0f}s / {max_wait_time}s)")
                time.sleep(5)
        except Exception as e:
            logger.error(f"Error al verificar el estado del archivo {file_id}: {e}", exc_info=True)
            raise 

def _manage_vector_store_async(new_file_id: str, old_file_id: str = None):
    with data_mentor_bp.app.app_context():
        try:
            logger.info(f"Iniciando proceso asíncrono para Vector Store con new_file_id: {new_file_id}")
            _wait_for_file_processing(new_file_id)
            
            vector_store_record = FileDailyID.query.first() 

            if vector_store_record and vector_store_record.current_vector_store_id:
                vector_store_id = vector_store_record.current_vector_store_id
                logger.info(f"Actualizando Vector Store existente: {vector_store_id} con el nuevo archivo {new_file_id}")
                
                vector_store_file = client.beta.vector_stores.files.create(
                    vector_store_id=vector_store_id,
                    file_id=new_file_id
                )
                logger.info(f"Archivo {new_file_id} adjuntado al Vector Store {vector_store_id}. Status de adjunción: {vector_store_file.status}")

                start_assimilation_time = time.time()
                max_assimilation_wait_time = 300 
                while True:
                    file_in_vector_store = client.beta.vector_stores.files.retrieve(
                        vector_store_id=vector_store_id,
                        file_id=new_file_id
                    )
                    if file_in_vector_store.status == "completed":
                        logger.info(f"Archivo {new_file_id} asimilado en Vector Store {vector_store_id} en {time.time() - start_assimilation_time:.2f} segundos.")
                        break
                    elif file_in_vector_store.status in ["failed", "cancelled"]:
                        error_msg = f"Fallo al asimilar archivo {new_file_id} en Vector Store {vector_store_id}. Estado: {file_in_vector_store.status}. Detalles: {file_in_vector_store.last_error.message if hasattr(file_in_vector_store, 'last_error') and file_in_vector_store.last_error else 'Desconocido'}"
                        logger.error(error_msg)
                        raise Exception(error_msg)
                    
                    current_assimilation_wait_time = time.time() - start_assimilation_time
                    if current_assimilation_wait_time > max_assimilation_wait_time:
                        error_msg = f"Tiempo de espera excedido ({max_assimilation_wait_time}s) para la asimilación del archivo {new_file_id} en Vector Store {vector_store_id}. Estado actual: {file_in_vector_store.status}"
                        logger.error(error_msg)
                        raise TimeoutError(error_msg)

                    logger.info(f"Archivo {new_file_id} en Vector Store {vector_store_id} en estado '{file_in_vector_store.status}', esperando asimilación... ({current_assimilation_wait_time:.0f}s / {max_assimilation_wait_time}s)")
                    time.sleep(5)

                if old_file_id and old_file_id != new_file_id:
                    try:
                        logger.info(f"Eliminando archivo antiguo {old_file_id} del Vector Store {vector_store_id}...")
                        client.beta.vector_stores.files.delete(
                            vector_store_id=vector_store_id,
                            file_id=old_file_id
                        )
                        logger.info(f"Archivo {old_file_id} eliminado exitosamente del Vector Store {vector_store_id}.")
                        try:
                            client.files.delete(old_file_id)
                            logger.info(f"Archivo antiguo '{old_file_id}' eliminado del storage de OpenAI (general).")
                        except Exception as e:
                            logger.warning(f"No se pudo eliminar el archivo antiguo '{old_file_id}' del storage global de OpenAI. Causa: {e}")
                    except Exception as e:
                        logger.warning(f"No se pudo eliminar el archivo antiguo {old_file_id} del Vector Store {vector_store_id}. Causa: {e}")

            else:
                logger.info("Creando un nuevo Vector Store para el nuevo archivo...")
                vector_store = client.beta.vector_stores.create(
                    name="Daily Knowledge Base",
                    file_ids=[new_file_id] 
                )
                vector_store_id = vector_store.id
                logger.info(f"Nuevo Vector Store creado. ID: {vector_store_id}. Archivo {new_file_id} adjuntado.")

                if not vector_store_record:
                    vector_store_record = FileDailyID(current_file_id=new_file_id, current_vector_store_id=vector_store_id)
                    db.session.add(vector_store_record)
                else:
                    vector_store_record.current_vector_store_id = vector_store_id
                    vector_store_record.current_file_id = new_file_id 
                db.session.commit()
                logger.info(f"Vector Store ID y File ID actualizados en la base de datos a: VS:{vector_store_id}, File:{new_file_id}")
            
            logger.info(f"Proceso asíncrono de gestión de Vector Store completado para el archivo {new_file_id}.")

        except Exception as e:
            db.session.rollback() 
            logger.error(f"Error fatal en el proceso asíncrono de gestión de Vector Store para archivo {new_file_id}: {e}", exc_info=True)

@data_mentor_bp.route("/actualizar-archivos-asistente", methods=["POST"])
def actualizar_archivos_asistente():
    start_time = datetime.now()
    tmpfile_path = None
    
    # Lista de modelos y sus nombres de sección en el JSON final
    # ASEGÚRATE de que estos nombres de modelo y atributos .serialize() sean correctos
    TABLES_TO_INCLUDE = [
        (Comentarios2025, "comentarios_2025", "total_registros"),
        (FichasGoogle, "fichas_google", "total_registros"),
        (FichasGoogleCompetencia, "fichas_google_competencia", "total_registros"),
        (Usuarios_Por_Asignacion, "usuarios_por_asignacion", "total_registros"),
        (Usuarios_Sin_ID, "usuarios_sin_id", "total_registros"),
        (ValidaUsuarios, "valida_usuarios", "total_registros"),
        (DetalleApies, "detalle_apies", "total_registros"),
        (AvanceCursada, "avance_cursada", "total_registros"),
        (DetallesDeCursos, "detalles_de_cursos", "total_registros"),
        (CursadasAgrupadas, "cursadas_agrupadas", "total_registros"),
        (FormularioGestor, "formulario_gestor", "total_registros"),
        (CuartoSurveySql, "cuarto_survey_sql", "total_registros"),
        (QuintoSurveySql, "quinto_survey_sql", "quinto_survey_sql"), # Ejemplo: si el nombre de la sección es diferente
        (Comentarios2023, "comentarios_2023", "total_registros"),
        (Comentarios2024, "comentarios_2024", "total_registros"),
        (BaseLoopEstaciones, "base_loop_estaciones", "total_registros"),
        (SalesForce, "sales_force", "total_registros"),
        (ComentariosCompetencia, "comentarios_competencia", "total_registros")
    ]
    
    # BATCH_SIZE para las consultas a la base de datos (para yield_per)
    DB_QUERY_BATCH_SIZE = 10000 

    logger.info("======================================================")
    logger.info("====== INICIANDO PROCESO DE ACTUALIZACIÓN DE ARCHIVO DE CONOCIMIENTO ======")
    logger.info("======================================================")
    logger.info(f"Hora de inicio: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("Recopilando datos de todas las tablas de forma eficiente...")

    content_summary = [] 
    resumen_conteos_totales = {} 

    try:
        with NamedTemporaryFile(mode="w+", delete=False, suffix=".json", encoding="utf-8") as tmpfile:
            tmpfile_path = tmpfile.name
            
            tmpfile.write('{\n')
            tmpfile.write('  "descripcion_contenido_archivo": "Este archivo JSON contiene datos operativos y de experiencia del cliente de YPF, organizados por sección. Cada sección incluye conteos y datos detallados. La sección \\"resumen_conteos_totales\\" provee acceso directo a los conteos por sección.",\n')
            
            logger.info("Recopilando conteos totales para el resumen...")
            for Model, section_name, _ in TABLES_TO_INCLUDE:
                count = db.session.query(Model).count()
                resumen_conteos_totales[section_name] = count
                logger.info(f"  - {section_name}: {count} registros.")
            
            # --- CORRECCIÓN DE SYNTAXERROR ---
            # Hacemos el json.dumps y los .replace en una variable separada
            # para evitar el backslash dentro del f-string directamente.
            conteos_json_str = json.dumps(resumen_conteos_totales, indent=2, ensure_ascii=False)
            # Ajustar la indentación para que los '{' y '}' de resumen_conteos_totales estén alineados
            conteos_json_str_formatted = conteos_json_str.replace("{\n ", "{\n    ").replace("\n}", "\n  }")
            tmpfile.write(f'  "resumen_conteos_totales": {conteos_json_str_formatted},\n')
            # --- FIN CORRECCIÓN ---

            first_section = True
            for Model, section_name, count_key in TABLES_TO_INCLUDE:
                if not first_section:
                    tmpfile.write(',\n') 
                else:
                    first_section = False

                logger.info(f"Procesando sección '{section_name}'...")
                
                tmpfile.write(f'  "{section_name}": {{\n')
                tmpfile.write(f'    "total_registros": {resumen_conteos_totales[section_name]},\n')
                tmpfile.write('    "datos": [\n')

                record_count_in_section = 0
                is_first_record_in_section = True

                offset = 0
                while True:
                    batch_of_records = db.session.query(Model).limit(DB_QUERY_BATCH_SIZE).offset(offset).all()
                    if not batch_of_records:
                        break 
                    
                    for item in batch_of_records:
                        if not is_first_record_in_section:
                            tmpfile.write(',\n')
                        else:
                            is_first_record_in_section = False
                        
                        json.dump(item.serialize(), tmpfile, indent=4, ensure_ascii=False)
                        record_count_in_section += 1
                    
                    offset += DB_QUERY_BATCH_SIZE
                    logger.info(f"  - '{section_name}': {record_count_in_section} registros procesados hasta ahora. Memoria liberada.")
                    
                    # db.session.expunge_all() # Desasocia objetos del lote para liberar memoria
                    # db.session.close() # Podría cerrar la conexión, Flask-SQLAlchemy lo maneja diferente.
                    # Si estás usando Flask-SQLAlchemy con Scoped Sessions, el `remove()` al final del request es mejor.
                    # Para forzar la liberación en un bucle largo:
                    db.session.rollback() # Limpia la sesión y libera objetos
                    # O si no quieres rollback para no perder estado de objetos no relacionados
                    # for obj in batch_of_records:
                    #     db.session.expunge(obj)
                    # del batch_of_records # Explicitamente borrar referencia al lote

                tmpfile.write('\n    ]\n') 
                tmpfile.write('  }') 
                
                content_summary.append({
                    "nombre": section_name,
                    "incluido": True, 
                    "peso_mb": "Calculado al final", 
                    "total_registros": record_count_in_section
                })
                logger.info(f"Sección '{section_name}' finalizada con {record_count_in_section} registros.")

            tmpfile.write('\n}') 
            tmpfile.flush()
            tmpfile.close() 
        
        file_size_bytes = os.path.getsize(tmpfile_path)
        file_size_mb = file_size_bytes / (1024 * 1024)
        logger.info(f"Tamaño final del archivo JSON temporal: {file_size_mb:.2f} MB ({file_size_bytes} bytes)")

        if file_size_bytes == 0:
            error_message = "El archivo JSON generado está vacío (0 bytes). No se puede subir a OpenAI para File Search."
            logger.error(error_message)
            return jsonify({
                "success": False,
                "message": error_message,
                "final_file_size_mb": 0,
                "contenido_incluido": content_summary
            }), 400

        # ... (El resto del código para subir a OpenAI y gestionar Vector Store es el mismo) ...
        # (Se mantiene la lógica de _manage_vector_store_async)
        logger.info("Subiendo el nuevo archivo JSON a OpenAI...")
        with open(tmpfile_path, "rb") as file_to_upload:
            uploaded_file = client.files.create(
                file=file_to_upload,
                purpose="assistants"
            )
        new_file_id = uploaded_file.id
        logger.info(f"Nuevo archivo JSON subido con éxito. File ID: {new_file_id}")

        existing_file_record = FileDailyID.query.first()
        old_file_id = None
        
        if existing_file_record:
            old_file_id = existing_file_record.current_file_id
            logger.info(f"Se encontró un registro existente en la DB. Old File ID: {old_file_id}")
            
            existing_file_record.current_file_id = new_file_id
            db.session.add(existing_file_record)
            db.session.commit()
            logger.info(f"ID de archivo actualizado en la base de datos a: {new_file_id}")
        else:
            logger.info("No se encontró un archivo anterior registrado en la base de datos. Creando nuevo registro.")
            new_record = FileDailyID(
                current_file_id=new_file_id,
                current_vector_store_id=None 
            )
            db.session.add(new_record)
            db.session.commit()
            logger.info(f"Nuevo registro de ID de archivo creado en la base de datos: {new_file_id}")
        
        executor.submit(_manage_vector_store_async, new_file_id, old_file_id)
        logger.info(f"Proceso de gestión de Vector Store iniciado en segundo plano para el archivo {new_file_id}.")

        end_time = datetime.now()
        time_elapsed = end_time - start_time
        seconds = int(time_elapsed.total_seconds())
        time_format = str(timedelta(seconds=seconds))

        response_message = {
            "success": True,
            "message": "Archivo de conocimiento diario recibido y proceso de actualización de OpenAI iniciado en segundo plano.",
            "new_file_id": new_file_id,
            "old_file_id_replaced": old_file_id if old_file_id else "N/A",
            "process_status": "El archivo se está subiendo y procesando. La creación/actualización del Vector Store se gestiona asíncronamente.",
            "final_file_size_mb": round(file_size_mb, 4),
            "tiempo_de_generacion_archivo_local": time_format, 
            "contenido_incluido": content_summary 
        }
        logger.info(f"Respuesta enviada al cliente. Detalles: {json.dumps(response_message, indent=2, ensure_ascii=False)}")
        return jsonify(response_message), 200

    except Exception as e:
        db.session.rollback() 
        end_time = datetime.now() 
        time_elapsed_on_error = end_time - start_time
        seconds_on_error = int(time_elapsed_on_error.total_seconds())
        time_format_on_error = str(timedelta(seconds=seconds_on_error))

        logger.error("======================================================")
        logger.error("============ ERROR FATAL DURANTE LA CARGA ============")
        logger.error(f"Hora del error: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        logger.error(f"Tiempo transcurrido hasta el error: {time_format_on_error}")
        logger.error(f"Error: {str(e)}", exc_info=True)
        logger.error("======================================================")
        return jsonify({"error": str(e), "status": 500}), 500
    finally:
        if tmpfile_path and os.path.exists(tmpfile_path):
            try:
                os.remove(tmpfile_path)
                logger.info(f"Archivo temporal '{tmpfile_path}' eliminado.")
            except PermissionError as pe:
                logger.error(f"Error de permiso al intentar eliminar el archivo temporal en el finally block: {pe}")
            except Exception as final_e:
                logger.error(f"Error inesperado al eliminar el archivo temporal en el finally block: {final_e}")


def _wait_for_file_processing(file_id: str):
    """Espera a que un archivo en OpenAI cambie su estado a 'processed'."""
    start_time = time.time()
    max_wait_time = 300 
    while True:
        try:
            file_obj = client.files.retrieve(file_id=file_id)
            if file_obj.status == "processed":
                logger.info(f"Archivo {file_id} procesado por OpenAI en {time.time() - start_time:.2f} segundos.")
                return file_obj
            elif file_obj.status == "failed":
                error_message = f"El procesamiento del archivo {file_id} falló en OpenAI. Detalles: {file_obj.last_error.message if file_obj.last_error else 'Desconocido'}."
                logger.error(error_message)
                raise Exception(error_message) 
            else:
                current_wait_time = time.time() - start_time
                if current_wait_time > max_wait_time:
                    error_message = f"Tiempo de espera excedido ({max_wait_time}s) para el procesamiento del archivo {file_id}. Estado actual: {file_obj.status}"
                    logger.error(error_message)
                    raise TimeoutError(error_message) 
                logger.info(f"Archivo {file_id} en estado '{file_obj.status}', esperando... ({current_wait_time:.0f}s / {max_wait_time}s)")
                time.sleep(5)
        except Exception as e:
            logger.error(f"Error al verificar el estado del archivo {file_id}: {e}", exc_info=True)
            raise 

def _wait_for_vector_store_completion(vector_store_id: str):
    """Espera a que un Vector Store termine de asimilar todos sus archivos."""
    start_time = time.time()
    max_wait_time = 600 # Aumentar el tiempo de espera para el Vector Store
    while True:
        try:
            vector_store = client.beta.vector_stores.retrieve(vector_store_id)
            # El estado 'completed' es la confirmación
            if vector_store.status == "completed":
                logger.info(f"Vector Store {vector_store_id} completado en {time.time() - start_time:.2f} segundos.")
                return vector_store
            elif vector_store.status in ["failed", "cancelled"]:
                error_message = f"La creación del Vector Store {vector_store_id} falló. Estado: {vector_store.status}."
                logger.error(error_message)
                raise Exception(error_message)
            else:
                current_wait_time = time.time() - start_time
                if current_wait_time > max_wait_time:
                    error_message = f"Tiempo de espera excedido ({max_wait_time}s) para la creación del Vector Store {vector_store_id}."
                    logger.error(error_message)
                    raise TimeoutError(error_message)
                logger.info(f"Vector Store {vector_store_id} en estado '{vector_store.status}', esperando... ({current_wait_time:.0f}s / {max_wait_time}s)")
                time.sleep(10) # Esperar un poco más para este proceso
        except Exception as e:
            logger.error(f"Error al verificar el estado del Vector Store {vector_store_id}: {e}", exc_info=True)
            raise

# -------------------------------------------------------------------------
# RUTA PRINCIPAL CON LA NUEVA LÓGICA DE SUBDIVISIÓN Y VECTORIZACIÓN
# -------------------------------------------------------------------------

@data_mentor_bp.route("/actualizar-comentarios-2025-subdividido", methods=["POST"])
def actualizar_comentarios_2025_subdividido():
    start_time = datetime.now()
    temp_dir = 'temp_openai_uploads'
    uploaded_file_ids = []
    vector_store_id = None

    # Parámetros para la creación de archivos
    MAX_FILE_SIZE_MB = 10
    DB_QUERY_BATCH_SIZE = 10000 
    
    logger.info("======================================================")
    logger.info("===== INICIANDO PROCESO SUBDIVIDIDO DE COMENTARIOS 2025 =====")
    logger.info("======================================================")
    logger.info(f"Hora de inicio: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"Tamaño máximo por archivo: {MAX_FILE_SIZE_MB} MB")

    try:
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)
            logger.info(f"Directorio temporal creado: {temp_dir}")
        
        # Paso 1: Generar archivos JSON subdivididos
        logger.info("Paso 1: Generando archivos JSON subdivididos...")
        
        Model = Comentarios2025
        count = db.session.query(Model).count()
        logger.info(f"Total de registros en Comentarios2025: {count}")
        
        file_part_number = 1
        current_file_path = None
        current_file_size_bytes = 0
        current_batch_of_records = []
        is_first_record_in_file = True
        
        offset = 0
        while True:
            batch_of_records = db.session.query(Model).limit(DB_QUERY_BATCH_SIZE).offset(offset).all()
            if not batch_of_records:
                # Escribir el lote final si existe
                if current_batch_of_records:
                    if not current_file_path: # Si no se ha abierto un archivo, abrir uno
                        current_file_path = os.path.join(temp_dir, f"comentarios_2025_parte_{file_part_number}.json")
                        with open(current_file_path, 'w', encoding='utf-8') as f:
                            f.write('{\n')
                            f.write(f'  "comentarios_2025_parte_{file_part_number}": [\n')
                            is_first_record_in_file = True
                            
                    with open(current_file_path, 'a', encoding='utf-8') as f:
                        for item in current_batch_of_records:
                            if not is_first_record_in_file:
                                f.write(',\n')
                            else:
                                is_first_record_in_file = False
                            json.dump(item.serialize(), f, indent=4, ensure_ascii=False)
                    
                    with open(current_file_path, 'a', encoding='utf-8') as f:
                        f.write('\n]\n')
                        f.write('}\n')
                    logger.info(f"Archivo final generado: {current_file_path}. Tamaño: {os.path.getsize(current_file_path) / (1024*1024):.2f} MB")
                break
            
            for item in batch_of_records:
                current_batch_of_records.append(item)
                
                # Simulación de tamaño del JSON para este registro
                # Esto es una estimación. Un cálculo más preciso sería serializar y medir.
                # Para un control estricto, hay que serializar y medir cada registro
                # pero eso puede ser lento. Este enfoque es más rápido.
                item_size_bytes_estimate = len(json.dumps(item.serialize(), ensure_ascii=False).encode('utf-8'))
                
                # Si el tamaño estimado del archivo actual + el nuevo registro supera el límite,
                # cerrar el archivo actual y comenzar uno nuevo.
                if current_file_path and (current_file_size_bytes + item_size_bytes_estimate) > (MAX_FILE_SIZE_MB * 1024 * 1024):
                    logger.info(f"Tamaño de archivo '{current_file_path}' ({current_file_size_bytes / (1024*1024):.2f} MB) alcanzado. Cerrando y comenzando nuevo.")
                    
                    with open(current_file_path, 'a', encoding='utf-8') as f:
                        f.write('\n]\n')
                        f.write('}\n')
                    
                    file_part_number += 1
                    current_file_path = None # Reset para abrir nuevo archivo
                    current_file_size_bytes = 0
                    current_batch_of_records = []
                    is_first_record_in_file = True

                # Si no hay un archivo abierto, crear uno nuevo
                if not current_file_path:
                    current_file_path = os.path.join(temp_dir, f"comentarios_2025_parte_{file_part_number}.json")
                    with open(current_file_path, 'w', encoding='utf-8') as f:
                        f.write('{\n')
                        f.write(f'  "comentarios_2025_parte_{file_part_number}": [\n')
                    is_first_record_in_file = True

                # Escribir el registro al archivo
                with open(current_file_path, 'a', encoding='utf-8') as f:
                    if not is_first_record_in_file:
                        f.write(',\n')
                    else:
                        is_first_record_in_file = False
                    json.dump(item.serialize(), f, indent=4, ensure_ascii=False)
                
                current_file_size_bytes = os.path.getsize(current_file_path)
            
            offset += DB_QUERY_BATCH_SIZE
            db.session.remove() # Limpiar la sesión para el siguiente lote
        
        # Finalizar el último archivo si no se ha hecho
        if current_file_path and not current_file_path.endswith('.json}'): # Verificar si ya se cerró el JSON
            with open(current_file_path, 'a', encoding='utf-8') as f:
                f.write('\n]\n')
                f.write('}\n')

        # Paso 2: Subir todos los archivos a OpenAI y esperar a que sean procesados
        logger.info("Paso 2: Subiendo y esperando a que todos los archivos sean procesados por OpenAI...")
        generated_files = [os.path.join(temp_dir, f) for f in os.listdir(temp_dir) if f.endswith('.json')]
        if not generated_files:
            raise Exception("No se generaron archivos para subir.")
        
        for file_path in generated_files:
            with open(file_path, "rb") as file_to_upload:
                uploaded_file = client.files.create(
                    file=file_to_upload,
                    purpose="assistants"
                )
            uploaded_file_ids.append(uploaded_file.id)
            logger.info(f"Archivo subido: '{os.path.basename(file_path)}', File ID: {uploaded_file.id}")
            _wait_for_file_processing(uploaded_file.id) # Esperar aquí de forma síncrona

        # Paso 3: Crear el Vector Store con todos los archivos
        logger.info("Paso 3: Todos los archivos procesados. Creando un único Vector Store...")
        
        vector_store = client.beta.vector_stores.create(
            name="Comentarios 2025 Knowledge Base",
            file_ids=uploaded_file_ids
        )
        vector_store_id = vector_store.id
        logger.info(f"Vector Store creado. ID: {vector_store_id}. Esperando a que se complete la asimilación...")
        _wait_for_vector_store_completion(vector_store_id) # Esperar a que el Vector Store esté listo
        
        # Paso 4: Actualizar la base de datos con el nuevo Vector Store ID
        logger.info(f"Paso 4: Actualizando la base de datos con el Vector Store ID: {vector_store_id}")
        existing_file_record = FileDailyID.query.first()
        old_file_id = existing_file_record.current_file_id if existing_file_record else None
        
        if existing_file_record:
            existing_file_record.current_file_id = ','.join(uploaded_file_ids) # Guardar todos los IDs
            existing_file_record.current_vector_store_id = vector_store_id
        else:
            new_record = FileDailyID(
                current_file_id=','.join(uploaded_file_ids),
                current_vector_store_id=vector_store_id
            )
            db.session.add(new_record)
            
        db.session.commit()
        logger.info(f"DB actualizada con los IDs de archivos y el nuevo Vector Store ID: {vector_store_id}.")
        
        end_time = datetime.now()
        time_elapsed = end_time - start_time
        seconds = int(time_elapsed.total_seconds())
        time_format = str(timedelta(seconds=seconds))
        
        logger.info("======================================================")
        logger.info("============ PROCESO FINALIZADO CON ÉXITO ============")
        logger.info(f"Tiempo total: {time_format}")
        logger.info("======================================================")

        return jsonify({
            "success": True,
            "message": f"Se crearon {len(uploaded_file_ids)} archivos y se generó un Vector Store único exitosamente.",
            "vector_store_id": vector_store_id,
            "archivos_creados": uploaded_file_ids,
            "tiempo_total": time_format
        }), 200

    except Exception as e:
        db.session.rollback() 
        logger.error(f"Error fatal durante el proceso: {str(e)}", exc_info=True)
        end_time = datetime.now()
        time_elapsed_on_error = end_time - start_time
        seconds_on_error = int(time_elapsed_on_error.total_seconds())
        time_format_on_error = str(timedelta(seconds=seconds_on_error))
        return jsonify({"error": str(e), "status": 500, "tiempo_transcurrido_hasta_error": time_format_on_error}), 500
    finally:
        if os.path.exists(temp_dir):
            for filename in os.listdir(temp_dir):
                file_path = os.path.join(temp_dir, filename)
                try:
                    if os.path.isfile(file_path):
                        os.remove(file_path)
                except Exception as e:
                    logger.error(f"Error al eliminar archivo temporal {file_path}: {e}")
            os.rmdir(temp_dir)
            logger.info(f"Directorio temporal '{temp_dir}' eliminado.")

