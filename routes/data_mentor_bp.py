from flask import Blueprint, send_file, make_response, request, jsonify, render_template, current_app, Response
from flask_bcrypt import Bcrypt 
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity 
from database import db 
from logging_config import logger
import os
from dotenv import load_dotenv 
load_dotenv()
from utils.data_mentor_utils import query_assistant_mentor
import urllib.request
import urllib.error
import json
import pandas as pd
from models import User,Instructions, ReportesDataMentor ,Usuarios_Por_Asignacion, Usuarios_Sin_ID, ValidaUsuarios,DetalleApies, AvanceCursada, DetallesDeCursos, CursadasAgrupadas,FormularioGestor,CuartoSurveySql, QuintoSurveySql, Comentarios2023, Comentarios2024, Comentarios2025, BaseLoopEstaciones, FichasGoogleCompetencia, FichasGoogle, SalesForce, ComentariosCompetencia, FileDailyID
from sqlalchemy.exc import SQLAlchemyError
import csv, textwrap
import time
import re
from openai import OpenAI, APIError, APITimeoutError
from tempfile import NamedTemporaryFile
import openpyxl
from io import BytesIO
from datetime import datetime, timedelta 
from sqlalchemy import create_engine, text
from sqlalchemy import inspect
from sqlalchemy.orm import sessionmaker
from typing import Dict, Any, List


OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")
OPENAI_MODEL = os.environ.get("OPENAI_MODEL", "gpt-4.1")
# OPENAI_MODEL= "gpt-3.5-turbo"
if not OPENAI_API_KEY:
    raise ValueError("Debes definir la variable de entorno OPENAI_API_KEY con tu clave de API.")

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

HEADERS = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {OPENAI_API_KEY}",
    "OpenAI-Beta": "assistants=v2"
}

data_mentor_bp = Blueprint('data_mentor_bp', __name__) 
bcrypt = Bcrypt()
jwt = JWTManager()

# Sistema de key base pre rutas ------------------------:

API_KEY = os.getenv('API_KEY')

def check_api_key(api_key):
    return api_key == API_KEY

@data_mentor_bp.before_request
def authorize():
    if request.method == 'OPTIONS':
        return
    if request.path in ['/horas-por-curso','/test_data_mentor_bp']:
        return
    api_key = request.headers.get('Authorization')
    if not api_key or not check_api_key(api_key):
        return jsonify({'message': 'Unauthorized'}), 401
    
# RUTA TEST:

VECTOR_STORE_ID = os.getenv("VECTOR_STORE_ID") 
DATABASE_URL = os.getenv("DATABASE_URL")

engine = create_engine(DATABASE_URL, pool_pre_ping=True)
Session = sessionmaker(bind=engine)

#Obtener NARRATIVA DE DB o usar la que ya estaba
def get_narrative_from_db():
    """Obtiene el texto de la instrucción más reciente desde la base de datos."""
    try:
        latest_instruction = Instructions.query.order_by(Instructions.created_at.desc()).first()
        if latest_instruction:
            return latest_instruction.instructions
        else:
            # Texto básico por si la tabla está vacía
            return textwrap.dedent("""
            Guía corta de datos y relaciones (narrativa, NO JSON)

            Idea general
            - La tabla principal es base_loop_estaciones. Tiene info operativa, geográfica y administrativa de cada estación.
            - Claves relevantes: "APIES" (identificador de estación, texto) y "Id" (otro identificador de estación).
            - Si un nombre de columna tiene espacios o signos (p. ej., Estacion de Servicio: Zona, Store Code), citá con comillas dobles: "Store Code".
            - Usá SQL ANSI. Strings con comillas simples. Identificadores con comillas dobles si tienen espacios.

            Relaciones típicas (joins)
            - base_loop_estaciones."APIES" = comentarios_encuesta_2023.apies
            - base_loop_estaciones."APIES" = comentarios_encuesta_2024.apies
            - base_loop_estaciones."APIES" = comentarios_encuesta_2025.apies

            - base_loop_estaciones."Id" = fichas_google."Store Code"
            - base_loop_estaciones."Id" = fichas_google_competencia.idLoop
            - base_loop_estaciones."Id" = comentarios_competencia.IDLOOP
            - base_loop_estaciones."Id" = usuarios_por_asignacion.id_pertenencia

            - **NUEVA:** La tabla detalle_apies contiene información de negocio. Su columna 'apies' se vincula a:
              - `usuarios_por_asignacion.id_pertenencia`
              - `base_loop_estaciones."Apies"`

            Aprendizaje / cursos (no todas tienen FK explícita, se relaciona por semántica):
            - dni aparece en varias: avance_cursada.dni, cursadas_agrupadas.dni, usuarios_* (dni).
            - id_curso en detalles_de_cursos.id_curso y cursadas_agrupadas.id_curso.
            - avance_cursada tiene nombre_corto_curso / nombre_programa (texto) y puede cruzar con detalles_de_cursos por nombre o id si existe un mapping conocido.

            Encuestas:
            - cuarto_survey_sql y quinto_survey_sql: resultados de encuestas (campos texto con preguntas). Vinculaciones por curso/gestor/fechas si hace falta.
            - comentarios_* (2023/2024/2025): comentarios de clientes por apies/fecha/sentiment.

            Competencia:
            - fichas_google_competencia e comentarios_competencia se enlazan a estaciones por idLoop ~ base_loop_estaciones."Id".
            - fichas_google: nuestras fichas (reseñas, rating) por "Store Code" ~ base_loop_estaciones."Id".

            Notas prácticas:
            - Preferí SELECT de columnas concretas (no *).
            - Si hay columnas con espacios, citá con "doble comilla".
            - Si pedís muchos registros, agregá LIMIT <= 200.
            """).strip()
    except Exception as e:
        # En caso de error de DB, usar el texto básico
        return f"Error al cargar la narrativa desde la base de datos: {str(e)}"
# ====== RELACIONES (NARRATIVA, SIN JSON) ======
NARRATIVE_RELATIONS = get_narrative_from_db()

# ====== Helpers: narrar el esquema sin JSON ======
def _table_summary(insp, table: str, max_cols: int | None = None) -> str:
    cols = [c["name"] for c in insp.get_columns(table)]
    if max_cols and len(cols) > max_cols:
        head = ", ".join(cols[:max_cols])
        return f"- {table}: columnas: {head}, … (+{len(cols)-max_cols} más)"
    else:
        return f"- {table}: columnas: {', '.join(cols) if cols else '(sin columnas detectadas)'}"

def build_db_schema_narrative(whitelist: List[str] | None = None, max_cols: int | None = None) -> str:
    """
    Devuelve un texto plano (no JSON) con el listado de tablas y sus columnas.
    Evita corchetes/llaves para no gatillar parsing raro.
    """
    insp = inspect(engine)
    tables = insp.get_table_names()
    if whitelist:
        tables = [t for t in tables if t in whitelist]

    lines = ["Esquema (resumen narrativo):"]
    for t in sorted(tables):
        try:
            lines.append(_table_summary(insp, t, max_cols=max_cols))
        except Exception:
            lines.append(f"- {t}: (no se pudieron listar columnas)")
    return "\n".join(lines)

# ====== Router ======
def make_router_system_prompt(narrative_schema: str) -> str:
    return textwrap.dedent(f"""
    Sos un router NL→(SQL|RAG|BOTH) para Data Mentor. Devolvés SOLO JSON válido (sin texto extra).

    REGLAS DURAS:
    - Elegí "SQL", "RAG" o "BOTH".
    - Si devolvés SQL: debe ser SOLO SELECT; sin INSERT/UPDATE/DELETE/DDL; sin ';', '--', '/* */'.
      Si falta LIMIT, agregalo (<= 200).
      Usá únicamente tablas y columnas mencionadas en el esquema narrativo.
      Si una columna tiene espacios o símbolos, citá con comillas dobles ("...").
      Preferí lista de columnas sobre SELECT *.
    - **NUEVA REGLA (FINAL):** Para buscar valores de texto en columnas, siempre usá `WHERE TRIM(LOWER(columna)) LIKE LOWER('%valor%')`. Para IDs o valores numéricos exactos, usá `=`.
    - Si no hay datos suficientes en DB o falta contexto textual → RAG_QUERY.
    - BOTH: incluye SELECT y RAG_QUERY.
    - REASON: 1–2 líneas cortas.

    CONTEXTO NARRATIVO (sin JSON):
    {NARRATIVE_RELATIONS}

    {narrative_schema}
    """).strip()

ROUTER_USER = """USER_TEXT:
{user_text}

Respondé SOLO con JSON:
{{
  "mode": "SQL" | "RAG" | "BOTH",
  "sql": "<SELECT ... o vacío>",
  "rag_query": "<texto o vacío>",
  "reason": "<1–2 líneas>"
}}"""

def call_router(user_text: str, narrative_schema_text: str) -> Dict[str, Any]:
    sys = make_router_system_prompt(narrative_schema_text)
    usr = ROUTER_USER.format(user_text=user_text)
    t0 = time.time()
    
    resp = client.chat.completions.create(
        model=OPENAI_MODEL,
        messages=[{"role":"system","content":sys}, {"role":"user","content":usr}],
        response_format={"type": "json_object"}
    )
    
    latency = time.time() - t0
    text_out = resp.choices[0].message.content.strip()
    out = json.loads(text_out)
    out["_router_latency"] = latency
    return out

# ====== SQL guard rails ======
FORBIDDEN = re.compile(r"\b(INSERT|UPDATE|DELETE|DROP|ALTER|TRUNCATE|MERGE|CREATE|GRANT|REVOKE)\b|;|--|/\*", re.I)

def _extract_tables(sql: str) -> List[str]:
    # match FROM/JOIN "Table Name" or from/join table_name (con/sin alias)
    pattern = re.compile(r'\b(from|join)\s+(".*?"|[\w\.]+)', re.I)
    tbls = []
    for m in pattern.finditer(sql):
        raw = m.group(2).strip()
        if raw.startswith('"') and raw.endswith('"'):
            raw = raw[1:-1]
        # quitar schema si viene schema.table
        raw = raw.split(".")[-1]
        tbls.append(raw)
    return list(dict.fromkeys(tbls)) 

def validate_sql(sql: str, allowed_tables: List[str]) -> str:
    s = sql.strip()
    if not s.lower().startswith("select"):
        raise ValueError("Solo SELECT permitido.")
    if FORBIDDEN.search(s):
        raise ValueError("Query contiene operaciones prohibidas o comentarios.")
    # chequeo de tablas usadas
    used = _extract_tables(s)
    for t in used:
        if t not in allowed_tables:
            raise ValueError(f"Tabla no permitida: {t}")
    # LIMIT <= 200
    m = re.search(r"\blimit\s+(\d+)", s, re.I)
    if m:
        n = int(m.group(1))
        if n > 200:
            s = re.sub(r"\blimit\s+\d+", "LIMIT 200", s, flags=re.I)
    else:
        s += " LIMIT 200"
    return s

def run_sql(sql: str) -> List[Dict[str, Any]]:
    session = Session()
    rows = session.execute(text(sql)).mappings().all()
    out = [dict(r) for r in rows]
    session.close()
    return out

# ====== Síntesis final (con RAG opcional) ======
def synthesize_final(user_text: str, results_sql: List[Dict[str,Any]] | None, rag_query: str | None) -> str:
    system = (
        "Sos Data Mentor. Respondés claro para managers.\n"
        "- Si hay resultados SQL: mostrás tabla corta (máx 10 filas) + resumen.\n"
        "- Si hay RAG: usá file_search (vector store adjunto) y citá fuente (archivo/sección).\n"
        "- Si falta info, decilo sin vueltas."
    )
    msgs = [
        {"role":"system","content":system},
        {"role":"user","content":(
            f"USER_TEXT: {user_text}\n\n"
            f"RESULTADOS_SQL (JSON, top 10): {json.dumps(results_sql[:10] if results_sql else [], ensure_ascii=False)}\n\n"
            f"RAG_QUERY: {rag_query or ''}\n"
            "Instrucciones:\n"
            "- Si RAG_QUERY no está vacío, ejecutá file_search para traer contexto y CITAR.\n"
            "- Devolvé respuesta final concisa con bullets + tabla si aplica."
        )}
    ]
    kwargs = {"model": OPENAI_MODEL, "messages": msgs}

    if rag_query:
        # El modelo de chat no necesita adjuntar herramientas como el Assistant.
        # Solo le damos la query para que sepa qué información buscar.
        # No se usa 'tool_resources' aquí.
        pass

    resp = client.chat.completions.create(**kwargs)
    return resp.choices[0].message.content


# ====== Ruta principal ======
@data_mentor_bp.route("/chat_mentor", methods=["POST"])
def chat_mentor():
    logger.info("1 - Entró en la ruta Chat_mentor.")
    data = request.get_json()
    if not data or "prompt" not in data:
        return jsonify({"error": "Falta el prompt en el cuerpo de la solicitud"}), 400

    user_text = data["prompt"]
    thread_id = data.get("thread_id")

    # 1) Armar guía narrativa del esquema (sin JSON).
    narrative_schema = build_db_schema_narrative(whitelist=None, max_cols=None)

    # 2) Router
    router = call_router(user_text, narrative_schema)
    mode = (router.get("mode") or "").upper()
    sql = (router.get("sql") or "").strip()
    rag_query = (router.get("rag_query") or "").strip()
    logger.info(f"Router => {mode} | reason={router.get('reason','')}")

    # 3) Ejecutar SQL (si aplica)
    resultados_sql: List[Dict[str, Any]] = []
    sql_ejecutado = ""
    try:
        if mode in ("SQL","BOTH") and sql:
            # allowed tables = las del schema narrado
            allowed_tables = inspect(engine).get_table_names()
            sql_validado = validate_sql(sql, allowed_tables)
            sql_ejecutado = sql_validado
            resultados_sql = run_sql(sql_validado)
    except Exception as e:
        return jsonify({"error": f"Error SQL: {str(e)}", "router": router}), 400

    # 4) Síntesis final (con RAG si aplica)
    try:
        final_text = synthesize_final(
            user_text,
            resultados_sql,
            rag_query if mode in ("RAG","BOTH") else None
        )
    except Exception as e:
        return jsonify({"error": f"Error en síntesis: {str(e)}", "router": router}), 500

    # 5) Respuesta
    return jsonify({
        "response": final_text,
        "thread_id": thread_id,
        "trace": {
            "mode": mode,
            "sql": sql_ejecutado,
            "rows": len(resultados_sql),
            "router_ms": int(router.get("_router_latency",0)*1000)
        }
    }), 200

@data_mentor_bp.route("/report_to_data_mentor", methods=["POST"])
def report_error_data_mentor():
    try:
        data = request.json
        if not data:
            return jsonify({"error": "No se recibieron datos JSON"}), 400

        user_email = data.get("user")
        if not user_email:
            return jsonify({"error": "El campo 'user' (email) es requerido"}), 400
        
        question = data.get("question")
        failed_answer = data.get("failed_answer")
        sql_query = data.get("sql_query")

        # Buscar el usuario por email para obtener el DNI
        user_record = User.query.filter_by(email=user_email).first()
        user_dni = user_record.dni if user_record else None

        # Crear una nueva instancia de ReportesDataMentor
        new_report = ReportesDataMentor(
            user=user_email,
            user_dni=user_dni,
            question=question,
            failed_answer=failed_answer,
            sql_query=sql_query,
            resolved=False
        )

        db.session.add(new_report)
        db.session.commit()

        return jsonify({"message": "Reporte de error guardado exitosamente"}), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@data_mentor_bp.route("/get_reports_of_data_mentor", methods=["GET"])
def get_reports_of_data_mentor():
    try:
        # Obtener todos los reportes de la base de datos
        reports = ReportesDataMentor.query.all()

        # Serializar la lista de reportes
        serialized_reports = [report.serialize() for report in reports]

        return jsonify(serialized_reports), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@data_mentor_bp.route("/get_instructions", methods=["GET"])
def get_instructions():
    try:
        # Obtener todas las instrucciones ordenadas por fecha de creación descendente
        instructions = Instructions.query.order_by(Instructions.created_at.desc()).all()
        serialized_instructions = [instr.serialize() for instr in instructions]
        return jsonify(serialized_instructions), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@data_mentor_bp.route("/set_instructions", methods=["POST"])
def set_instructions():
    try:
        data = request.json
        instructions_text = data.get("instructions")
        user_email = data.get("user")

        if not instructions_text or not user_email:
            return jsonify({"error": "Los campos 'instructions' y 'user' son requeridos."}), 400

        # Crear un nuevo registro de instrucciones
        new_instructions = Instructions(
            user=user_email,
            instructions=instructions_text
        )
        db.session.add(new_instructions)
        db.session.commit()

        return jsonify({"message": "Instrucciones guardadas exitosamente."}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@data_mentor_bp.route("/delete_instructions", methods=["DELETE"])
def delete_instructions():
    try:
        data = request.json
        instruction_id = data.get("id")

        if not instruction_id:
            return jsonify({"error": "El campo 'id' es requerido."}), 400

        instruction = Instructions.query.get(instruction_id)
        if not instruction:
            return jsonify({"error": "Registro de instrucciones no encontrado."}), 404

        db.session.delete(instruction)
        db.session.commit()

        return jsonify({"message": f"Instrucción con ID {instruction_id} eliminada exitosamente."}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


#TODOESTO ES PARA QUE FUNCIONE EL "REPARAR CON IA"

TABLES_WHITELIST = [
    'Usuarios_Por_Asignacion', 'Usuarios_Sin_ID', 'ValidaUsuarios', 'DetalleApies',
    'AvanceCursada', 'DetallesDeCursos', 'CursadasAgrupadas', 'FormularioGestor',
    'CuartoSurveySql', 'QuintoSurveySql', 'Comentarios2023', 'Comentarios2024',
    'Comentarios2025', 'BaseLoopEstaciones', 'FichasGoogleCompetencia',
    'FichasGoogle', 'SalesForce', 'ComentariosCompetencia',
]

def _table_summary(insp, table: str, max_cols: int | None = None) -> str:
    try:
        cols = [c["name"] for c in insp.get_columns(table)]
        if max_cols and len(cols) > max_cols:
            head = ", ".join(cols[:max_cols])
            return f"- {table}: columnas: {head}, … (+{len(cols)-max_cols} más)"
        else:
            return f"- {table}: columnas: {', '.join(cols) if cols else '(sin columnas detectadas)'}"
    except Exception:
        return f"- {table}: (no se pudieron listar columnas)"

def build_db_schema_narrative(whitelist: List[str] | None = None, max_cols: int | None = None) -> str:
    engine = db.engine
    insp = inspect(engine)
    tables = insp.get_table_names()
    if whitelist:
        tables = [t for t in tables if t in whitelist]

    lines = ["Esquema (resumen narrativo):"]
    for t in sorted(tables):
        lines.append(_table_summary(insp, t, max_cols=max_cols))
    return "\n".join(lines)


@data_mentor_bp.route("/fix_instructions_by_error", methods=["POST"])
def fix_instructions_by_error():
    if client is None:
        error_msg = "Error de configuración del servidor: La clave de la API de OpenAI no es válida."
        logger.error(error_msg)
        return jsonify({"error": error_msg}), 500

    try:
        data = request.get_json()
        report_id = data.get("id")

        if not report_id:
            logger.error("ERROR: Falta el ID del reporte.")
            return jsonify({"error": "Falta el ID del reporte."}), 400

        logger.info(f"DEBUG: Procesando el reporte con ID: {report_id}")

        reporte = ReportesDataMentor.query.get(report_id)
        if not reporte:
            logger.error(f"ERROR: Reporte con ID {report_id} no encontrado.")
            return jsonify({"error": "Reporte no encontrado."}), 404

        logger.info("DEBUG: Reporte encontrado. Obteniendo las instrucciones actuales...")

        instrucciones_actuales = Instructions.query.order_by(Instructions.created_at.desc()).first()
        if not instrucciones_actuales:
            logger.error("ERROR: No hay instrucciones de IA disponibles.")
            return jsonify({"error": "No hay instrucciones de IA disponibles."}), 500

        logger.info("DEBUG: Instrucciones actuales encontradas. Obteniendo el esquema de tablas...")
        esquema_tablas = build_db_schema_narrative(whitelist=TABLES_WHITELIST)
        logger.info("DEBUG: Esquema de tablas generado.")

        # --- PROMPT MEJORADO Y MÁS EXPLÍCITO SOBRE EL ORDEN ---
        prompt_template = textwrap.dedent("""
        Analiza el siguiente error de un sistema de IA que genera SQL. Tienes la tarea de mejorar las instrucciones que guían a esa IA para que no cometa el mismo error en el futuro.

        **REGLA CRÍTICA:** Al mejorar las instrucciones, asegúrate de **mantener la totalidad del contenido original** y solo añadir o modificar lo necesario para corregir el error. La coherencia del sistema depende de mantener las instrucciones previas.

        Instrucciones actuales:
        {instrucciones_actuales}

        Detalles del error:
        - Pregunta del usuario: "{pregunta_usuario}"
        - Respuesta fallida de la IA: "{respuesta_fallida}"
        - SQL incorrecto utilizado: "{sql_utilizado}"

        Contexto para la mejora:
        - Estructura de las tablas:
        {esquema_tablas}

        Si las instrucciones actuales son adecuadas y el error no se debe a ellas, por favor devuélvelas sin cambios y explica el motivo.

        Tu respuesta debe tener **exactamente** el siguiente formato textual, sin ningún texto adicional. Asegúrate de que **NUEVA_INSTRUCCION** sea el primer campo y que **MOTIVO_EXPLICACION** lo siga inmediatamente.

        NUEVA_INSTRUCCION:"{nueva_instruccion}"
        MOTIVO_EXPLICACION:"{motivo_de_los_cambios}"
        """)

        llm_prompt = prompt_template.format(
            instrucciones_actuales=instrucciones_actuales.instructions,
            pregunta_usuario=reporte.question,
            respuesta_fallida=reporte.failed_answer,
            sql_utilizado=reporte.sql_query if reporte.sql_query else "No se utilizó SQL.",
            esquema_tablas=esquema_tablas,
            nueva_instruccion="",
            motivo_de_los_cambios=""
        )
        
        logger.info(f"DEBUG: Tamaño del prompt a enviar: %s caracteres.", len(llm_prompt))
        logger.info("DEBUG: Prompt para el LLM construido. Llamando a la API de OpenAI...")

        messages = [
            {"role": "system", "content": "Eres un asistente experto en optimización de instrucciones para modelos de lenguaje. Tu única tarea es analizar un error y proponer una nueva instrucción mejorada. Si las instrucciones son correctas, devuélvelas sin cambios y explícame por qué."},
            {"role": "user", "content": llm_prompt}
        ]
        kwargs = {
            "model": OPENAI_MODEL,
            "messages": messages,
            "timeout": 120.0
        }

        try:
            t0 = time.time()
            response_llm = client.chat.completions.create(**kwargs)
            llm_latency = time.time() - t0
            logger.info(f"DEBUG: Respuesta del LLM recibida en {llm_latency:.2f} segundos.")
            llm_text_out = response_llm.choices[0].message.content
            logger.info(f"DEBUG: Respuesta cruda del LLM:\n{llm_text_out}")
            
            # --- Lógica de parsing con la nueva regex ---
            nueva_instruccion_match = re.search(r'NUEVA_INSTRUCCION:"(.*?)"\s*MOTIVO_EXPLICACION:', llm_text_out, re.DOTALL)
            motivo_explicacion_match = re.search(r'MOTIVO_EXPLICACION:"(.*?)"', llm_text_out, re.DOTALL)
            
            if nueva_instruccion_match:
                nueva_instruccion = nueva_instruccion_match.group(1).strip()
            else:
                nueva_instruccion = ""
            
            if motivo_explicacion_match:
                motivo_explicacion = motivo_explicacion_match.group(1).strip()
            else:
                motivo_explicacion = ""

        except APITimeoutError:
            error_msg = "La API de OpenAI excedió el tiempo de espera. Por favor, intenta de nuevo."
            logger.error(f"ERROR: Fallo de la API de OpenAI por timeout. Mensaje: {error_msg}")
            return jsonify({"error": error_msg}), 500
        except APIError as api_error:
            logger.error(f"ERROR: Fallo de la API de OpenAI: %s", api_error.response.text)
            return jsonify({"error": f"Fallo de la API de OpenAI: {api_error.response.text}"}), 500
        except Exception as e:
            logger.error("ERROR: Error inesperado al procesar la respuesta del LLM: %s", str(e))
            return jsonify({"error": f"Error al procesar la respuesta de la IA: {str(e)}"}), 500

        if not nueva_instruccion or not motivo_explicacion:
            logger.error("ERROR: El LLM devolvió un formato incorrecto y no se pudo extraer la instrucción o el motivo.")
            return jsonify({"error": "El LLM no devolvió los campos esperados."}), 500

        logger.info("DEBUG: Respuesta del LLM parseada con éxito.")
        reporte.resolved = True
        db.session.commit()
        logger.info("DEBUG: Reporte %s marcado como resuelto.", report_id)

        logger.info("DEBUG: Enviando respuesta al frontend.")
        return jsonify({
            "nueva_instruccion": nueva_instruccion,
            "motivo_explicacion": motivo_explicacion
        }), 200

    except Exception as e:
        db.session.rollback()
        logger.error("ERROR: Fallo inesperado en fix_instructions_by_error: %s", str(e))
        return jsonify({"error": f"Fallo inesperado: {str(e)}"}), 500
#-------------------------------------------------


@data_mentor_bp.route("/switch_error_status", methods=["POST"])
def switch_error_status():
    try:
        data = request.get_json()
        report_id = data.get("id")

        if not report_id:
            logger.error("ERROR: Falta el ID del reporte en la solicitud.")
            return jsonify({"error": "Falta el ID del reporte."}), 400

        reporte = ReportesDataMentor.query.get(report_id)
        if not reporte:
            logger.error(f"ERROR: Reporte con ID {report_id} no encontrado.")
            return jsonify({"error": "Reporte no encontrado."}), 404

        # Cambiar el estado de resolved
        reporte.resolved = not reporte.resolved
        db.session.commit()

        logger.info(f"DEBUG: El estado del reporte {report_id} se ha cambiado a {reporte.resolved}.")
        return jsonify({
            "message": "Estado del reporte actualizado con éxito.",
            "new_status": reporte.resolved,
        }), 200

    except Exception as e:
        db.session.rollback()
        logger.error(f"ERROR: Fallo inesperado al cambiar el estado del reporte: {str(e)}")
        return jsonify({"error": f"Fallo inesperado: {str(e)}"}), 500



@data_mentor_bp.route("/close_chat_mentor", methods=["POST"])
def close_chat():
    """
    Ruta para cerrar el thread del chat.
    Se espera recibir un JSON con la clave "thread_id".
    Llama al endpoint DELETE de la API para cerrar el hilo usando urllib.
    """
    data = request.get_json()
    if not data or "thread_id" not in data:
        return jsonify({"error": "Falta el thread_id en el cuerpo de la solicitud"}), 400

    thread_id = data["thread_id"]
    delete_url = f"https://api.openai.com/v1/threads/{thread_id}"

    try:
        req = urllib.request.Request(delete_url, headers=HEADERS, method="DELETE")
        with urllib.request.urlopen(req) as response:
            result_data = response.read().decode("utf-8")
            result = json.loads(result_data)
        return jsonify(result), 200
    except urllib.error.HTTPError as e:
        error_message = e.read().decode("utf-8")
        return jsonify({"error": f"HTTPError {e.code}: {error_message}"}), e.code
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@data_mentor_bp.route('/horas-por-curso', methods=['GET'])
def horas_por_curso():
    data = [
        {"curso": "Node.js Básico", "horas": 5},
        {"curso": "React Intermedio", "horas": 7},
        {"curso": "Flask Fullstack", "horas": 9}
    ]
    return jsonify(data)

# -------------------------- MODELOS QUE SE TIENEN EN CUENTA PARA CONTABILIZAR SUS REGISTROS ----------------------------------

# Diccionario para mapear nombres de string a clases reales
MODELS = {
    'Usuarios_Por_Asignacion': Usuarios_Por_Asignacion,
    'Usuarios_Sin_ID': Usuarios_Sin_ID,
    'ValidaUsuarios': ValidaUsuarios,
    'DetalleApies' : DetalleApies,
    'AvanceCursada': AvanceCursada,
    'DetallesDeCursos' : DetallesDeCursos,
    'CursadasAgrupadas' : CursadasAgrupadas,
    'FormularioGestor' :FormularioGestor,
    'CuartoSurveySql': CuartoSurveySql,
    'QuintoSurveySql' : QuintoSurveySql,
    'Comentarios2023': Comentarios2023,
    'Comentarios2024': Comentarios2024,
    'Comentarios2025': Comentarios2025,
    'BaseLoopEstaciones' : BaseLoopEstaciones,
    'FichasGoogleCompetencia' : FichasGoogleCompetencia,
    'FichasGoogle' : FichasGoogle,
    'SalesForce' : SalesForce,
    'ComentariosCompetencia' : ComentariosCompetencia
    # Agregá los modelos que quieras habilitar acá
}

# -------------------------- Contabilizar longitud de cualquier tabla ----------------------------------

@data_mentor_bp.route('/contar-registros', methods=['POST'])
def contar_registros():
    data = request.get_json()
    nombre_tabla = data.get('tabla')

    if not nombre_tabla:
        return jsonify({"error": "Falta el nombre de la tabla"}), 400

    modelo = MODELS.get(nombre_tabla)
    if not modelo:
        return jsonify({"error": f"La tabla '{nombre_tabla}' no está habilitada o no existe"}), 404

    try:
        cantidad = db.session.query(modelo).count()
        return jsonify({"tabla": nombre_tabla, "cantidad_registros": cantidad}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# -------------------------- ACA VIENEN LAS RUTAS DE LAS TABLAS DE REPORTES ----------------------------


@data_mentor_bp.route('/usuarios_por_asignacion/<int:registro_id>', methods=['GET'])
def get_usuario_por_asignacion(registro_id):
    """
    Devuelve el registro de Usuarios_Por_Asignacion con el id dado,
    usando el método serialize() del modelo.
    """
    logger.info(f"Buscando Usuarios_Por_Asignacion id={registro_id}")
    registro = Usuarios_Por_Asignacion.query.get(registro_id)

    if not registro:
        logger.warning(f"Usuarios_Por_Asignacion id={registro_id} no encontrado")
        return jsonify({
            'error': 'Registro no encontrado',
            'status': 404
        }), 404

    logger.info(f"Registro encontrado: {registro}")
    return jsonify(registro.serialize()), 200

@data_mentor_bp.route('/usuarios_sin_id/<int:registro_id>', methods=['GET'])
def get_usuario_sin_id(registro_id):
    """
    Devuelve el registro de Usuarios_Sin_ID con el id dado,
    usando el método serialize() del modelo.
    """
    logger.info(f"Buscando Usuarios_Sin_ID id={registro_id}")
    registro = Usuarios_Sin_ID.query.get(registro_id)

    if not registro:
        logger.warning(f"Usuarios_Sin_ID id={registro_id} no encontrado")
        return jsonify({
            'error': 'Registro no encontrado',
            'status': 404
        }), 404

    logger.info(f"Registro encontrado: {registro}")
    return jsonify(registro.serialize()), 200




# RUTAS PARA CARGAR TABLAS DE EXPERIENCIA 2023 24 y 25

@data_mentor_bp.route('/cargar_comentarios_2023', methods=['POST'])
def cargar_comentarios_encuesta_2023():
    """
    Recibe un archivo .xlsx vía form-data (campo: 'file') y guarda sus registros en la DB
    """
    archivo = request.files.get('file')
    if not archivo:
        return jsonify({'error': 'No se envió ningún archivo', 'status': 400}), 400

    try:
        df = pd.read_excel(archivo)

        registros = []
        for _, fila in df.iterrows():
            fecha_raw = fila.get('FECHA')
            try:
                fecha = pd.to_datetime(fecha_raw) if pd.notnull(fecha_raw) else None
            except:
                fecha = None

            nuevo = Comentarios2023(
                fecha=fecha,
                apies=str(fila.get('APIES', '')).strip(),
                comentario=str(fila.get('COMENTARIO', '')).strip(),
                canal=str(fila.get('CANAL', '')).strip(),
                topico=str(fila.get('TÓPICO', '')).strip(),
                sentiment=str(fila.get('SENTIMENT', '')).strip()
            )
            registros.append(nuevo)

        db.session.add_all(registros)
        db.session.commit()

        return jsonify({'mensaje': f'Se guardaron {len(registros)} comentarios', 'status': 200}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}', 'status': 500}), 500
    
@data_mentor_bp.route('/cargar_comentarios_2024', methods=['POST'])
def cargar_comentarios_encuesta_2024():
    """
    Recibe un archivo .xlsx vía form-data (campo: 'file') y guarda sus registros en la DB.
    Optimizado para bajo uso de memoria mediante procesamiento fila a fila con openpyxl.
    Incluye logs de progreso consolidados y tiempo de ejecución en la respuesta.
    """
    start_time = datetime.now() # Iniciar el temporizador

    archivo = request.files.get('file')
    if not archivo:
        logger.error("No se envió ningún archivo en la solicitud.")
        return jsonify({'error': 'No se envió ningún archivo', 'status': 400}), 400

    total_registros_guardados = 0
    BATCH_SIZE = 5000 
    
    # Contadores y muestras para errores de fecha
    date_errors_count = 0
    sample_date_errors = []
    MAX_SAMPLE_DATE_ERRORS = 5 # Cuántos ejemplos de errores de fecha guardar

    logger.info("======================================================")
    logger.info("============= INICIANDO PROCESO DE CARGA ==============")
    logger.info("======================================================")
    logger.info(f"Hora de inicio: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"Archivo recibido: '{archivo.filename}' ({archivo.content_length / (1024*1024):.2f} MB)")
    logger.info(f"Tamaño de lote (BATCH_SIZE) para DB: {BATCH_SIZE} filas por commit.")

    try:
        excel_data = BytesIO(archivo.read())
        workbook = openpyxl.load_workbook(excel_data, read_only=True, data_only=True)
        
        if not workbook.sheetnames:
            logger.error("El archivo Excel recibido no contiene hojas.")
            return jsonify({'error': 'El archivo Excel no contiene hojas.', 'status': 400}), 400
        
        sheet = workbook[workbook.sheetnames[0]] 
        logger.info(f"Se procesará la hoja: '{sheet.title}'")

        # Asumimos que la primera fila son los encabezados
        headers = [cell.value for cell in sheet[1]] 
        header_map = {str(header).strip().upper(): i for i, header in enumerate(headers) if header}
        
        col_fecha = header_map.get('FECHA')
        col_apies = header_map.get('APIES')
        col_comentario = header_map.get('COMENTARIO')
        col_canal = header_map.get('CANAL')
        col_topico = header_map.get('TÓPICO')
        col_sentiment = header_map.get('SENTIMENT')

        if col_fecha is None or col_apies is None or col_comentario is None or \
           col_canal is None or col_topico is None or col_sentiment is None:
            missing_cols = []
            if col_fecha is None: missing_cols.append('FECHA')
            if col_apies is None: missing_cols.append('APIES')
            if col_comentario is None: missing_cols.append('COMENTARIO')
            if col_canal is None: missing_cols.append('CANAL')
            if col_topico is None: missing_cols.append('TÓPICO')
            if col_sentiment is None: missing_cols.append('SENTIMENT')
            
            error_msg = f"Faltan columnas requeridas en el archivo Excel: {', '.join(missing_cols)}. Asegúrate de que los encabezados coincidan (MAYÚSCULAS)."
            logger.error(error_msg)
            return jsonify({'error': error_msg, 'status': 400}), 400

        registros_chunk = []
        row_counter_total = 0 
        chunk_counter = 0

        # Iterar sobre las filas de la hoja, empezando desde la segunda fila
        # enumerate(..., start=2) para obtener el número de fila real del Excel (incluyendo encabezado)
        for row_index_excel, row in enumerate(sheet.iter_rows(min_row=2), start=2): 
            row_counter_total += 1
            
            fecha_raw = row[col_fecha].value if col_fecha is not None else None
            apies = row[col_apies].value if col_apies is not None else ''
            comentario = row[col_comentario].value if col_comentario is not None else ''
            canal = row[col_canal].value if col_canal is not None else ''
            topico = row[col_topico].value if col_topico is not None else ''
            sentiment = row[col_sentiment].value if col_sentiment is not None else ''
            
            fecha = None # Inicializar fecha como None
            try:
                if isinstance(fecha_raw, datetime):
                    fecha = fecha_raw # openpyxl ya lo parseó a datetime
                elif pd.notnull(fecha_raw) and str(fecha_raw).strip(): # Si tiene valor no nulo y no es string vacío
                    # Intentar parsear si no es datetime
                    fecha = pd.to_datetime(fecha_raw, errors='coerce') 
                # Si pd.to_datetime con errors='coerce' falla, fecha será NaT, que se convierte en None al asignar
            except Exception as date_e:
                # Solo logear si es un error que no sea solo 'coerce' a NaT, o si queremos más detalle
                pass # No loguear cada error de fecha individualmente
            
            # Contar errores de fecha y guardar una muestra
            if fecha is pd.NaT: # pd.NaT indica que pd.to_datetime falló y coecionó
                date_errors_count += 1
                if len(sample_date_errors) < MAX_SAMPLE_DATE_ERRORS:
                    sample_date_errors.append(f"Fila {row_index_excel}: '{fecha_raw}'")
                fecha = None # Asegurarse de que sea None para la DB si falló el parseo

            nuevo = Comentarios2024(
                fecha=fecha,
                apies=str(apies).strip(),
                comentario=str(comentario).strip(),
                canal=str(canal).strip(),
                topico=str(topico).strip(),
                sentiment=str(sentiment).strip()
            )
            registros_chunk.append(nuevo)

            # Si el chunk alcanza el tamaño definido, lo guardamos en la DB
            if len(registros_chunk) >= BATCH_SIZE:
                chunk_counter += 1
                logger.info(f"--- Procesando CHUNK {chunk_counter} (Filas Excel {row_counter_total - len(registros_chunk) + 1} - {row_counter_total}) ---")
                
                db.session.add_all(registros_chunk)
                db.session.commit()
                total_registros_guardados += len(registros_chunk)
                logger.info(f"CHUNK {chunk_counter} COMPLETO. Guardados {len(registros_chunk)} registros en DB. Total acumulado: {total_registros_guardados}")
                
                db.session.expunge_all() 
                del registros_chunk 
                registros_chunk = [] # Reiniciar la lista para el siguiente lote
                
                logger.info(f"--- Memoria del CHUNK {chunk_counter} liberada. ---")

        # Guardar los registros restantes (el último lote)
        if registros_chunk:
            chunk_counter += 1
            logger.info(f"--- Procesando CHUNK FINAL {chunk_counter} (Filas Excel {row_counter_total - len(registros_chunk) + 1} - {row_counter_total}) ---")
            db.session.add_all(registros_chunk)
            db.session.commit()
            total_registros_guardados += len(registros_chunk)
            logger.info(f"CHUNK FINAL {chunk_counter} COMPLETO. Guardados {len(registros_chunk)} registros en DB.")
            logger.info(f"TOTAL DE REGISTROS GUARDADOS EN LA DB: {total_registros_guardados}")
            db.session.expunge_all()
            del registros_chunk

        end_time = datetime.now() # Finalizar el temporizador
        time_elapsed = end_time - start_time
        # Formatear el tiempo de ejecución a HH:MM:SS
        seconds = int(time_elapsed.total_seconds())
        time_format = str(timedelta(seconds=seconds))

        logger.info("======================================================")
        logger.info("============ PROCESO DE CARGA FINALIZADO ============")
        logger.info(f"Hora de finalización: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"Tiempo total de ejecución: {time_format}")
        logger.info(f"TOTAL DE FILAS PROCESADAS EN EXCEL: {row_counter_total}")
        logger.info(f"TOTAL DE REGISTROS GUARDADOS EN LA DB: {total_registros_guardados}")
        if date_errors_count > 0:
            logger.warning(f"ADVERTENCIA: Se encontraron {date_errors_count} errores al parsear fechas. Ejemplos: {'; '.join(sample_date_errors)}")
        logger.info("======================================================")
        
        return jsonify({
            'mensaje': f'Se guardaron {total_registros_guardados} comentarios en total.',
            'status': 200,
            'tiempo_de_guardado': time_format,
            'detalles_procesamiento': {
                'total_filas_excel_leidas': row_counter_total,
                'total_registros_db_guardados': total_registros_guardados,
                'errores_fecha_contados': date_errors_count,
                'ejemplos_errores_fecha': sample_date_errors if date_errors_count > 0 else "Ninguno"
            }
        }), 200

    except Exception as e:
        db.session.rollback() 
        end_time = datetime.now() # Capturar tiempo de finalización incluso en error
        time_elapsed_on_error = end_time - start_time
        seconds_on_error = int(time_elapsed_on_error.total_seconds())
        time_format_on_error = str(timedelta(seconds=seconds_on_error))

        logger.error("======================================================")
        logger.error("============ ERROR FATAL DURANTE LA CARGA ============")
        logger.error(f"Hora del error: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        logger.error(f"Tiempo transcurrido hasta el error: {time_format_on_error}")
        logger.error(f"Error: {str(e)}", exc_info=True)
        logger.error("======================================================")
        return jsonify({
            'error': f'Error al procesar el archivo: {str(e)}', 
            'status': 500,
            'tiempo_transcurrido_hasta_error': time_format_on_error
        }), 500
    
@data_mentor_bp.route('/cargar_comentarios_2025', methods=['POST'])
def cargar_comentarios_encuesta_2025():
    """
    Recibe un archivo .xlsx vía form-data (campo: 'file') y guarda sus registros en la DB.
    Implementa lógica incremental (deduplicación por hash_id) y optimización de memoria
    mediante openpyxl y procesamiento por lotes.
    Incluye logs detallados de progreso y tiempo de ejecución.
    """
    start_time = datetime.now() # Iniciar el temporizador

    archivo = request.files.get('file')
    if not archivo:
        logger.error("No se envió ningún archivo en la solicitud.")
        return jsonify({'error': 'No se envió ningún archivo', 'status': 400}), 400

    # Contadores para el resumen final
    total_registros_guardados = 0
    total_filas_excel_leidas = 0
    date_errors_count = 0
    sample_date_errors = []
    MAX_SAMPLE_DATE_ERRORS = 5 
    duplicados_en_archivo_contados = 0
    duplicados_en_db_contados = 0

    # Este set rastreará todos los hashes únicos encontrados en el ARCHIVO ACTUAL
    # para evitar duplicados que aparezcan en diferentes lotes del mismo archivo.
    all_hashes_seen_in_current_file_session = set() 

    BATCH_SIZE = 5000 

    logger.info("======================================================")
    logger.info("====== INICIANDO PROCESO DE CARGA COMENTARIOS 2025 ======")
    logger.info("======================================================")
    logger.info(f"Hora de inicio: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"Archivo recibido: '{archivo.filename}' ({archivo.content_length / (1024*1024):.2f} MB)")
    logger.info(f"Tamaño de lote (BATCH_SIZE) para DB: {BATCH_SIZE} filas por commit.")

    try:
        excel_data = BytesIO(archivo.read())
        workbook = openpyxl.load_workbook(excel_data, read_only=True, data_only=True)
        
        if not workbook.sheetnames:
            logger.error("El archivo Excel recibido no contiene hojas.")
            return jsonify({'error': 'El archivo Excel no contiene hojas.', 'status': 400}), 400
        
        sheet = workbook[workbook.sheetnames[0]] 
        logger.info(f"Se procesará la hoja: '{sheet.title}'")

        # Asumimos que la primera fila son los encabezados
        headers = [cell.value for cell in sheet[1]] 
        header_map = {str(header).strip().upper(): i for i, header in enumerate(headers) if header}
        
        # Mapeo de columnas, asegurando que existan
        col_fecha = header_map.get('FECHA')
        col_apies = header_map.get('APIES')
        col_comentario = header_map.get('COMENTARIO')
        col_canal = header_map.get('CANAL')
        col_topico = header_map.get('TÓPICO')
        col_sentiment = header_map.get('SENTIMENT')

        # Verificar que las columnas obligatorias existan
        required_cols = {'FECHA', 'APIES', 'COMENTARIO', 'CANAL', 'TÓPICO', 'SENTIMENT'}
        missing_cols = [col for col in required_cols if header_map.get(col) is None]
        if missing_cols:
            error_msg = f"Faltan columnas requeridas en el archivo Excel: {', '.join(missing_cols)}. Asegúrate de que los encabezados coincidan (MAYÚSCULAS)."
            logger.error(error_msg)
            return jsonify({'error': error_msg, 'status': 400}), 400

        registros_para_lote_db = [] # Esta lista contendrá los objetos ORM para el bulk_save_objects
        chunk_counter = 0

        # Iterar sobre las filas de la hoja, empezando desde la segunda fila (después del encabezado)
        for row_index_excel, row in enumerate(sheet.iter_rows(min_row=2), start=2): 
            total_filas_excel_leidas += 1
            
            # Extraer valores de celda de forma segura
            fecha_raw = row[col_fecha].value
            apies = row[col_apies].value
            comentario = row[col_comentario].value
            canal = row[col_canal].value
            topico = row[col_topico].value
            sentiment = row[col_sentiment].value
            
            # --- Manejo de la fecha y generación del hash ---
            fecha = None
            try:
                if isinstance(fecha_raw, datetime):
                    fecha = fecha_raw 
                elif pd.notnull(fecha_raw) and str(fecha_raw).strip():
                    fecha = pd.to_datetime(fecha_raw, errors='coerce') 
                
                if fecha is pd.NaT: # Si el parseo de Pandas falló y coecionó
                    date_errors_count += 1
                    if len(sample_date_errors) < MAX_SAMPLE_DATE_ERRORS:
                        sample_date_errors.append(f"Fila {row_index_excel}: '{fecha_raw}'")
                    fecha = None # Asegurarse de que sea None para la DB si falló el parseo
            except Exception as date_e:
                date_errors_count += 1
                if len(sample_date_errors) < MAX_SAMPLE_DATE_ERRORS:
                    sample_date_errors.append(f"Fila {row_index_excel}: '{fecha_raw}' - {date_e}")
                fecha = None 
            
            # Stripear strings antes de generar hash y crear objeto
            apies = str(apies if apies is not None else '').strip()
            comentario = str(comentario if comentario is not None else '').strip()
            canal = str(canal if canal is not None else '').strip()
            topico = str(topico if topico is not None else '').strip()
            sentiment = str(sentiment if sentiment is not None else '').strip()

            # Generar hash único para este registro
            hash_id = Comentarios2025.generar_hash(fecha, apies, comentario, canal, topico, sentiment)

            # Verificar si este hash_id ya fue visto en el archivo ACTUAL (en cualquier lote procesado)
            if hash_id in all_hashes_seen_in_current_file_session:
                duplicados_en_archivo_contados += 1
                # logger.debug(f"Saltando duplicado DENTRO DEL ARCHIVO (hash: {hash_id}) en fila {row_index_excel}.")
                continue # Saltar esta fila, ya fue procesada o es un duplicado interno

            all_hashes_seen_in_current_file_session.add(hash_id) # Registrar este hash_id como visto

            comentario_obj = Comentarios2025(
                fecha=fecha,
                apies=apies,
                comentario=comentario,
                canal=canal,
                topico=topico,
                sentiment=sentiment,
                hash_id=hash_id # Asignar el hash_id al objeto
            )
            registros_para_lote_db.append(comentario_obj)

            # Si el lote para DB alcanza el tamaño definido, procesamos y guardamos
            if len(registros_para_lote_db) >= BATCH_SIZE:
                chunk_counter += 1
                logger.info(f"--- Procesando LOTE {chunk_counter} ({len(registros_para_lote_db)} candidatos para DB) ---")
                
                # Obtener los hashes de este lote para consultar la DB
                hashes_en_lote = [obj.hash_id for obj in registros_para_lote_db]
                
                # Consultar la DB para ver cuáles de estos hashes ya existen
                existentes_en_db_lote = set(
                    r[0] for r in db.session.query(Comentarios2025.hash_id)
                    .filter(Comentarios2025.hash_id.in_(hashes_en_lote))
                    .all()
                )
                
                # Filtrar los que realmente son nuevos para insertar
                nuevos_para_insertar = [
                    obj for obj in registros_para_lote_db 
                    if obj.hash_id not in existentes_en_db_lote
                ]
                
                duplicados_en_db_contados += (len(registros_para_lote_db) - len(nuevos_para_insertar))

                if nuevos_para_insertar:
                    # Usamos bulk_save_objects para una inserción más eficiente
                    db.session.bulk_save_objects(nuevos_para_insertar)
                    db.session.commit()
                    total_registros_guardados += len(nuevos_para_insertar)
                    logger.info(f"LOTE {chunk_counter} COMPLETO. Insertados {len(nuevos_para_insertar)} nuevos registros en DB.")
                    logger.info(f"TOTAL ACUMULADO GUARDADOS: {total_registros_guardados}")
                    logger.info(f"TOTAL ACUMULADO DUPLICADOS EN DB IGNORADOS: {duplicados_en_db_contados}")
                else:
                    logger.info(f"LOTE {chunk_counter} COMPLETO. No se encontraron registros nuevos para insertar en DB.")
                
                db.session.expunge_all() 
                del registros_para_lote_db 
                registros_para_lote_db = [] # Reiniciar la lista para el siguiente lote
                
                logger.info(f"--- Memoria del LOTE {chunk_counter} liberada. ---")

        # --- Fin del bucle de filas ---

        # Procesar y guardar los registros restantes (el último lote, si lo hay)
        if registros_para_lote_db:
            chunk_counter += 1
            logger.info(f"--- Procesando LOTE FINAL {chunk_counter} ({len(registros_para_lote_db)} candidatos restantes) ---")
            
            hashes_en_lote = [obj.hash_id for obj in registros_para_lote_db]
            existentes_en_db_lote = set(
                r[0] for r in db.session.query(Comentarios2025.hash_id)
                .filter(Comentarios2025.hash_id.in_(hashes_en_lote))
                .all()
            )
            nuevos_para_insertar = [
                obj for obj in registros_para_lote_db 
                if obj.hash_id not in existentes_en_db_lote
            ]
            duplicados_en_db_contados += (len(registros_para_lote_db) - len(nuevos_para_insertar))

            if nuevos_para_insertar:
                db.session.bulk_save_objects(nuevos_para_insertar)
                db.session.commit()
                total_registros_guardados += len(nuevos_para_insertar)
                logger.info(f"LOTE FINAL {chunk_counter} COMPLETO. Insertados {len(nuevos_para_insertar)} nuevos registros en DB.")
            else:
                logger.info(f"LOTE FINAL {chunk_counter} COMPLETO. No se encontraron registros nuevos para insertar en DB.")
            
            db.session.expunge_all()
            del registros_para_lote_db

        end_time = datetime.now() 
        time_elapsed = end_time - start_time
        seconds = int(time_elapsed.total_seconds())
        time_format = str(timedelta(seconds=seconds))

        logger.info("======================================================")
        logger.info("========== PROCESO DE CARGA COMENTARIOS 2025 FINALIZADO ==========")
        logger.info(f"Hora de finalización: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"Tiempo total de ejecución: {time_format}")
        logger.info(f"TOTAL DE FILAS LEÍDAS DEL EXCEL: {total_filas_excel_leidas}")
        logger.info(f"TOTAL DE REGISTROS NUEVOS GUARDADOS EN LA DB: {total_registros_guardados}")
        logger.info(f"TOTAL DE DUPLICADOS EN EL ARCHIVO IGNORADOS: {duplicados_en_archivo_contados}")
        logger.info(f"TOTAL DE DUPLICADOS EN LA DB IGNORADOS: {duplicados_en_db_contados}")
        if date_errors_count > 0:
            logger.warning(f"ADVERTENCIA: Se encontraron {date_errors_count} errores al parsear fechas. Ejemplos: {'; '.join(sample_date_errors)}")
        logger.info("======================================================")
        
        return jsonify({
            'mensaje': f'Se guardaron {total_registros_guardados} comentarios nuevos en total.',
            'status': 200,
            'tiempo_de_guardado': time_format,
            'detalles_procesamiento': {
                'total_filas_excel_leidas': total_filas_excel_leidas,
                'total_registros_db_guardados': total_registros_guardados,
                'duplicados_ignorados_en_archivo': duplicados_en_archivo_contados,
                'duplicados_ignorados_en_db': duplicados_en_db_contados,
                'errores_fecha_contados': date_errors_count,
                'ejemplos_errores_fecha': sample_date_errors if date_errors_count > 0 else "Ninguno"
            }
        }), 200

    except Exception as e:
        db.session.rollback() 
        end_time = datetime.now() 
        time_elapsed_on_error = end_time - start_time
        seconds_on_error = int(time_elapsed_on_error.total_seconds())
        time_format_on_error = str(timedelta(seconds=seconds_on_error))

        logger.error("======================================================")
        logger.error("============ ERROR FATAL DURANTE LA CARGA ============")
        logger.error(f"Hora del error: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        logger.error(f"Tiempo transcurrido hasta el error: {time_format_on_error}")
        logger.error(f"Error: {str(e)}", exc_info=True)
        logger.error("======================================================")
        return jsonify({
            'error': f'Error al procesar el archivo: {str(e)}', 
            'status': 500,
            'tiempo_transcurrido_hasta_error': time_format_on_error,
            'detalles_procesamiento': {
                'total_filas_excel_leidas': total_filas_excel_leidas,
                'total_registros_db_guardados_hasta_error': total_registros_guardados # Registros guardados antes del error
            }
        }), 500
    
@data_mentor_bp.route('/cargar_base_loop', methods=['POST'])
def cargar_base_loop():
    archivo = request.files.get('file')
    if not archivo:
        return jsonify({"error": "No se envió ningún archivo"}), 400

    try:
        # Detectar delimitador (coma o punto y coma)
        sample = archivo.read(2048).decode('utf-8')
        archivo.seek(0)
        delimiter = csv.Sniffer().sniff(sample).delimiter

        # Leer el archivo CSV de forma optimizada
        df = pd.read_csv(archivo, sep=delimiter, encoding='utf-8', on_bad_lines='skip')
    except Exception as e:
        return jsonify({"error": f"Error al leer el archivo CSV: {e}"}), 400

    try:
        # CORRECCIÓN DEFINITIVA:
        # 1. Crear un diccionario que mapee el nombre de la columna del CSV (el alias)
        #    al nombre del atributo que usa el modelo de SQLAlchemy.
        columnas_csv_a_modelo_attr = {
            'Id': 'id',
            'Apies': 'apies',
            'Inscripcion': 'inscripcion',
            'Operador': 'operador',
            'Estado Boca': 'estado_boca',
            'Bandera': 'bandera',
            'Direccion Admin.': 'direccion_admin',
            'Localidad Geo.': 'localidad_geo',
            'Provincia Geo.': 'provincia_geo',
            'Region Geo.': 'region_geo',
            'Zona Com. Geo.': 'zona_com_geo',
            'Tipo Establecimiento': 'tipo_establecimiento',
            'Tipo Operador': 'tipo_operador',
            'Tipo Despacho': 'tipo_despacho',
            'Tipo Ubicacion': 'tipo_ubicacion',
            'Region Admin.': 'region_admin',
            'Zona Com. Admin.': 'zona_com_admin',
            'RRCC': 'rrcc',
            'Opera ACA': 'opera_aca',
            'Bandera ACA': 'bandera_aca',
            'Contrato GN': 'contrato_gn',
            'Tipo Operacion': 'tipo_operacion',
            'Inicio actividad cuenta': 'inicio_actividad_cuenta',
            'Tipo Imagen': 'tipo_imagen',
            'Tipo Tienda': 'tipo_tienda',
            'Tipo Lubricentro': 'tipo_lubricentro',
            'Serviclub': 'serviclub',
            'YPF Ruta': 'ypf_ruta',
            'Azul 32': 'azul_32',
            'Punto Eléctrico': 'punto_eléctrico',
            'Cant. Imagenes': 'cant_imagenes',
            'Cant. Surtidores': 'cant_surtidores',
            'Cant. Tanques': 'cant_tanques',
            'Cant. Bombas': 'cant_bombas',
            'Cant. Conectores': 'cant_conectores',
            'Vol. prom. total N2 (m3)': 'vol_prom_total_n2_m3',
            'Vol. prom. total N3 (m3)': 'vol_prom_total_n3_m3',
            'Vol. prom. total Nafta (m3)': 'vol_prom_total_nafta_m3',
            'Vol. prom. total GO (m3)': 'vol_prom_total_go_m3',
            'Vol. prom. total GO2 (m3)': 'vol_prom_total_go2_m3',
            'Vol. prom. total GO3 (m3)': 'vol_prom_total_go3_m3',
            'Volumen Promedio Liquidos (m3)': 'volumen_promedio_liquidos_m3',
            'Vol. prom. GNC (m3)': 'vol_prom_gnc_m3',
            'Volumen Promedio Lubricantes (m3)': 'volumen_promedio_lubricantes_m3',
            'Cantidad Cambios de Lubricantes': 'cantidad_cambios_de_lubricantes',
            'Facturacion Bruta Promedio Tienda (ARS)': 'facturacion_bruta_promedio_tienda_ars',
            'Despacho_Liq_Prom_N2': 'despacho_liq_prom_n2',
            'Despacho_Liq_Prom_N3': 'despacho_liq_prom_n3',
            'Despacho_Liq_Prom_Nafta': 'despacho_liq_prom_nafta',
            'Despacho_Liq_Prom_GO2': 'despacho_liq_prom_go2',
            'Despacho_Liq_Prom_GO3': 'despacho_liq_prom_go3',
            'Despacho_Liq_Prom_Gasoil': 'despacho_liq_prom_gasoil',
            'Despacho_Liq_Prom_Total': 'despacho_liq_prom_total',
            'Despacho_Cnt_Prom_N3': 'despacho_cnt_prom_n3',
            'Despacho_Cnt_Prom_Nafta': 'despacho_cnt_prom_nafta',
            'Despacho_Cnt_Prom_GO2': 'despacho_cnt_prom_go2',
            'Despacho_Cnt_Prom_GO3': 'despacho_cnt_prom_go3',
            'Despacho_Cnt_Prom_Gasoil': 'despacho_cnt_prom_gasoil',
            'Despacho_Cnt_Prom_Total': 'despacho_cnt_prom_total',
            'Despacho_Vol_Prom_N2': 'despacho_vol_prom_n2',
            'Despacho_Vol_Prom_N3': 'despacho_vol_prom_n3',
            'Despacho_Vol_Prom_Nafta': 'despacho_vol_prom_nafta',
            'Despacho_Vol_Prom_GO2': 'despacho_vol_prom_go2',
            'Despacho_Vol_Prom_GO3': 'despacho_vol_prom_go3',
            'Despacho_Vol_Prom_Gasoil': 'despacho_vol_prom_gasoil',
            'Despacho_Vol_Prom_Total': 'despacho_vol_prom_total',
            'YPF Ruta Credito_Vol_Prom_N2': 'ypf_ruta_credito_vol_prom_n2',
            'YPF Ruta Credito_Vol_Prom_N3': 'ypf_ruta_credito_vol_prom_n3',
            'YPF Ruta Credito_Vol_Prom_Nafta': 'ypf_ruta_credito_vol_prom_nafta',
            'YPF Ruta Credito_Vol_Prom_GO2': 'ypf_ruta_credito_vol_prom_go2',
            'YPF Ruta Credito_Vol_Prom_GO3': 'ypf_ruta_credito_vol_prom_go3',
            'YPF Ruta Credito_Vol_Prom_Gasoil': 'ypf_ruta_credito_vol_prom_gasoil',
            'YPF Ruta Credito_Vol_Prom_Total': 'ypf_ruta_credito_vol_prom_total',
            'YPF Ruta Contado_Vol_Prom_N2': 'ypf_ruta_contado_vol_prom_n2',
            'YPF Ruta Contado_Vol_Prom_N3': 'ypf_ruta_contado_vol_prom_n3',
            'YPF Ruta Contado_Vol_Prom_Nafta': 'ypf_ruta_contado_vol_prom_nafta',
            'YPF Ruta Contado_Vol_Prom_GO2': 'ypf_ruta_contado_vol_prom_go2',
            'YPF Ruta Contado_Vol_Prom_GO3': 'ypf_ruta_contado_vol_prom_go3',
            'YPF Ruta Contado_Vol_Prom_Gasoil': 'ypf_ruta_contado_vol_prom_gasoil',
            'YPF Ruta Contado_Vol_Prom_Total': 'ypf_ruta_contado_vol_prom_total',
            'Serviclub_Penetracion_Por_N2': 'serviclub_penetracion_por_n2',
            'Serviclub_Penetracion_Por_N3': 'serviclub_penetracion_por_n3',
            'Serviclub_Penetracion_Por_Nafta': 'serviclub_penetracion_por_nafta',
            'Serviclub_Penetracion_Por_GO2': 'serviclub_penetracion_por_go2',
            'Serviclub_Penetracion_Por_GO3': 'serviclub_penetracion_por_go3',
            'Serviclub_Penetracion_Por_Gasoil': 'serviclub_penetracion_por_gasoil',
            'Serviclub_Penetracion_Por_Total': 'serviclub_penetracion_por_total',
            'Serviclub_Vol_Base_N2': 'serviclub_vol_base_n2',
            'Serviclub_Vol_Base_N3': 'serviclub_vol_base_n3',
            'Serviclub_Vol_Base_Nafta': 'serviclub_vol_base_nafta',
            'Serviclub_Vol_Base_GO2': 'serviclub_vol_base_go2',
            'Serviclub_Vol_Base_GO3': 'serviclub_vol_base_go3',
            'Serviclub_Vol_Base_Gasoil': 'serviclub_vol_base_gasoil',
            'Serviclub_Vol_Base_Total': 'serviclub_vol_base_total',
            'Dotación Actual_Total': 'dotacion_actual_total',
            'Dotación Actual_Jefes de Estación': 'dotacion_actual_jefes_de_estacion',
            'Dotación Actual_Jefes Trainee': 'dotacion_actual_jefes_trainee',
            'Dotación Actual_Responsables de Turno': 'dotacion_actual_responsables_de_turno',
            'Dotación Actual_Vendedor Dual': 'dotacion_actual_vendedor_dual',
            'Dotación Actual_Vendedor SR': 'dotacion_actual_vendedor_sr',
            'Dotación Actual_Lubriexperto': 'dotacion_actual_lubriexperto',
            'Dotación Actual_Lubriplaya': 'dotacion_actual_lubriplaya',
            'Descripción Tramo 1': 'descripcion_tramo_1',
            'Porcentaje Tramo 1': 'porcentaje_tramo_1',
            'Descripción Tramo 2': 'descripcion_tramo_2',
            'Porcentaje Tramo 2': 'porcentaje_tramo_2',
            'CUIT': 'cuit',
            'Red Propia': 'red_propia',
            'Zona Exclusión': 'zona_exclusion',
            'Nivel Socio Económico': 'nivel_socio_economico',
            'Densidad Poblacional (Hab/km2)': 'densidad_poblacional_hab_por_km2',
            'Latitud': 'latitud',
            'Longitud': 'longitud',
        }
        
        registros = []
        for _, fila in df.iterrows():
            datos_instancia = {}
            for nombre_col_csv, valor in fila.items():
                nombre_col_csv = nombre_col_csv.strip()
                
                # Usamos el mapeo para obtener el nombre del atributo de Python
                nombre_atributo_modelo = columnas_csv_a_modelo_attr.get(nombre_col_csv)
                
                if nombre_atributo_modelo:
                    if pd.isna(valor):
                        datos_instancia[nombre_atributo_modelo] = None
                    else:
                        datos_instancia[nombre_atributo_modelo] = valor

            # Instanciamos el modelo usando el diccionario con los nombres de atributos correctos
            instancia = BaseLoopEstaciones(**datos_instancia)
            registros.append(instancia)

        db.session.bulk_save_objects(registros)
        db.session.commit()

        return jsonify({"mensaje": f"{len(registros)} registros insertados correctamente"}), 200

    except SQLAlchemyError as e:
        db.session.rollback()
        logger.error(f"Error en la base de datos: {e}", exc_info=True)
        return jsonify({"error": f"Error al insertar en la base de datos: {e}"}), 500
    except Exception as e:
        logger.error(f"Error al procesar el archivo: {e}", exc_info=True)
        return jsonify({"error": f"Error al procesar el archivo: {e}"}), 500
    
@data_mentor_bp.route('/cargar_fichas_google_competencia', methods=['POST'])
def cargar_fichas_google_competencia():
    archivo = request.files.get('file')
    if not archivo:
        return jsonify({'error': 'No se envió ningún archivo', 'status': 400}), 400

    try:
        df = pd.read_excel(archivo)

        candidatos = []
        hash_ids = []
        hash_set_memoria = set()

        for _, fila in df.iterrows():
            id_loop = str(fila.get('idLoop', '')).strip()
            total_review_count = str(fila.get('totalReviewCount', '')).strip()
            average_rating = str(fila.get('averageRating', '')).strip()

            # Saltear filas con campos vacíos importantes
            if not id_loop or not total_review_count or not average_rating:
                continue

            hash_id = FichasGoogleCompetencia.generar_hash(id_loop, total_review_count, average_rating)

            # Saltear si ya está en memoria (repetido en el mismo archivo)
            if hash_id in hash_set_memoria:
                continue
            hash_set_memoria.add(hash_id)

            ficha_obj = FichasGoogleCompetencia(
                id_loop=id_loop,
                total_review_count=total_review_count,
                average_rating=average_rating,
                hash_id=hash_id
            )

            candidatos.append(ficha_obj)
            hash_ids.append(hash_id)

        # Buscar duplicados ya existentes en la DB
        existentes = set(
            r[0] for r in db.session.query(FichasGoogleCompetencia.hash_id)
            .filter(FichasGoogleCompetencia.hash_id.in_(hash_ids))
            .all()
        )

        nuevos = [f for f in candidatos if f.hash_id not in existentes]

        if nuevos:
            db.session.bulk_save_objects(nuevos)
            db.session.commit()

        return jsonify({
            'mensaje': f'Se guardaron {len(nuevos)} fichas nuevas',
            'preexistentes_ignorados': len(candidatos) - len(nuevos),
            'status': 200
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}', 'status': 500}), 500

@data_mentor_bp.route('/cargar_fichas_google', methods=['POST'])
def cargar_fichas_google():
    archivo = request.files.get('file')
    if not archivo:
        return jsonify({'error': 'No se envió ningún archivo', 'status': 400}), 400

    try:
        df = pd.read_excel(archivo)

        candidatos = []
        hash_ids = []
        hash_set_memoria = set()

        for _, fila in df.iterrows():
            store_code = str(fila.get('Store Code', '')).strip()
            cantidad_de_calificaciones = str(fila.get('Cantidad de calificaciones', '')).strip()
            start_rating = str(fila.get('Star Rating', '')).strip()

            # Saltear filas con campos vacíos importantes
            if not store_code or not cantidad_de_calificaciones or not start_rating:
                continue

            hash_id = FichasGoogle.generar_hash(store_code, cantidad_de_calificaciones, start_rating)

            # Saltear si ya está en memoria (repetido en el mismo archivo)
            if hash_id in hash_set_memoria:
                continue
            hash_set_memoria.add(hash_id)

            ficha_obj = FichasGoogle(
                store_code=store_code,
                cantidad_de_calificaciones=cantidad_de_calificaciones,
                start_rating=start_rating,
                hash_id=hash_id
            )

            candidatos.append(ficha_obj)
            hash_ids.append(hash_id)

        # Buscar duplicados ya existentes en la DB
        existentes = set(
            r[0] for r in db.session.query(FichasGoogle.hash_id)
            .filter(FichasGoogle.hash_id.in_(hash_ids))
            .all()
        )

        nuevos = [f for f in candidatos if f.hash_id not in existentes]

        if nuevos:
            db.session.bulk_save_objects(nuevos)
            db.session.commit()

        return jsonify({
            'mensaje': f'Se guardaron {len(nuevos)} fichas nuevas',
            'preexistentes_ignorados': len(candidatos) - len(nuevos),
            'status': 200
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}', 'status': 500}), 500
    
@data_mentor_bp.route('/cargar_salesforce', methods=['POST'])
def cargar_salesforce():
    archivo = request.files.get('file')
    if not archivo:
        return jsonify({'error': 'No se envió ningún archivo', 'status': 400}), 400

    try:
        df = pd.read_excel(archivo)

        candidatos = []
        hash_ids = []
        hash_set_memoria = set()

        for _, fila in df.iterrows():
            valores = [
                fila.get('Estacion de Servicio: Zona', ''),
                fila.get('Número del caso', ''),
                fila.get('Estado', ''),
                fila.get('Tipificación Caso', ''),
                fila.get('Asunto', ''),
                fila.get('Fecha/Hora de apertura', ''),
                fila.get('Cantidad de Reclamos', ''),
                fila.get('Defensa al Consumidor', ''),
                fila.get('GGRR/COLA Asignado', ''),
                fila.get('Propietario del caso: Nombre completo', ''),
                fila.get('Descripción', ''),
                fila.get('Nombre del contacto: Nombre completo', ''),
                fila.get('Comentarios', ''),
                fila.get('Estacion de Servicio: Razón Social', ''),
                fila.get('Estacion de Servicio: Red', ''),
                fila.get('Estacion de Servicio: Regional', ''),
            ]

            hash_id = SalesForce.generar_hash(*valores)

            if hash_id in hash_set_memoria:
                continue
            hash_set_memoria.add(hash_id)

            registro = SalesForce(
                estacion_servicio_zona=valores[0],
                numero_de_caso=valores[1],
                estado=valores[2],
                tipificacion_caso=valores[3],
                asunto=valores[4],
                fecha_apertura=valores[5],
                cantidad_reclamos=valores[6],
                defensa_consumidor=valores[7],
                ggrr_cola_asignado=valores[8],
                propietario_nombre=valores[9],
                descripcion=valores[10],
                contacto_nombre=valores[11],
                comentarios=valores[12],
                razon_social=valores[13],
                red=valores[14],
                regional=valores[15],
                hash_id=hash_id
            )

            candidatos.append(registro)
            hash_ids.append(hash_id)

        existentes = set(
            r[0] for r in db.session.query(SalesForce.hash_id)
            .filter(SalesForce.hash_id.in_(hash_ids))
            .all()
        )

        nuevos = [r for r in candidatos if r.hash_id not in existentes]

        if nuevos:
            db.session.bulk_save_objects(nuevos)
            db.session.commit()

        return jsonify({
            'mensaje': f'Se guardaron {len(nuevos)} casos nuevos',
            'preexistentes_ignorados': len(candidatos) - len(nuevos),
            'status': 200
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}', 'status': 500}), 500
    
@data_mentor_bp.route('/cargar_comentarios_competencia', methods=['POST'])
def cargar_comentarios_competencia():
    archivo = request.files.get('file')
    if not archivo:
        return jsonify({'error': 'No se recibió ningún archivo'}), 400

    try:
        df = pd.read_excel(archivo)

        # Normalizar la fecha
        if 'FECHA' in df.columns:
            df['FECHA'] = df['FECHA'].astype(str)


        # Renombrar columna ID si existe
        if 'ID' in df.columns:
            df.rename(columns={'ID': 'ID_ORIGINAL'}, inplace=True)

        # Reemplazamos espacios y ponemos mayúsculas para asegurar
        df.columns = [col.upper().replace(" ", "_") for col in df.columns]

        nuevos = []
        for _, fila in df.iterrows():
            hash_id = ComentariosCompetencia.generar_hash(
                fila.get('ID_ORIGINAL', ''),
                fila.get('FECHA', ''),
                fila.get('IDLOOP', ''),
                fila.get('COMENTARIO', ''),
                fila.get('RATING', ''),
                fila.get('SENTIMIENTO', ''),
                fila.get('TÓPICO', '')
            )

            # Chequeamos si ya existe
            if not ComentariosCompetencia.query.filter_by(hash_id=hash_id).first():
                nuevo = ComentariosCompetencia(
                    id_original=fila.get('ID_ORIGINAL'),
                    fecha=fila.get('FECHA'),
                    id_loop=fila.get('IDLOOP'),
                    comentario=fila.get('COMENTARIO'),
                    rating=fila.get('RATING'),
                    sentimiento=fila.get('SENTIMIENTO'),
                    topico=fila.get('TÓPICO'),
                    hash_id=hash_id
                )
                nuevos.append(nuevo)

        db.session.bulk_save_objects(nuevos)
        db.session.commit()

        return jsonify({'guardados': len(nuevos)})

    except Exception as e:
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}'}), 500
    

